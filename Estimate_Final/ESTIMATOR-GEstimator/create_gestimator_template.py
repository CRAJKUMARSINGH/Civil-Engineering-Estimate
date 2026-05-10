#!/usr/bin/env python3
"""
GEstimator Template Generator
Creates Excel templates compatible with GEstimator import format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import argparse

def create_schedule_template(filename="gestimator_schedule_template.xlsx"):
    """Create a template for Schedule Items"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule_Template"
    
    # Headers
    headers = ['Code', 'Description', 'Unit', 'Rate', 'Qty', 'Amount', 'Remarks']
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add sample data
    sample_data = [
        ['1', 'Earth work excavation in foundation trenches', 'Cum', 178, 720, 128160, 'All kinds of soil'],
        ['2', 'Concrete work M25 grade', 'Cum', 5003, 34.8, 174104, 'Including curing'],
        ['3', 'Brick work in superstructure', 'Sqm', 299, 1250, 373750, 'FPS bricks'],
        ['4', 'Steel reinforcement work', 'Quintal', 7650, 45, 344250, 'TMT bars'],
        ['5', 'Plastering internal walls', 'Sqm', 87, 2400, 208800, '15mm thick'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Set column widths
    column_widths = [12, 50, 8, 12, 12, 15, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Add data validation for Unit column
    unit_validation = DataValidation(
        type="list",
        formula1='"Cum,Sqm,Rmt,Nos,Kg,Quintal,MT,Ltr,Each,Job,LS"',
        allow_blank=True
    )
    unit_validation.error = "Please select from the dropdown list"
    unit_validation.errorTitle = "Invalid Unit"
    ws.add_data_validation(unit_validation)
    unit_validation.add(f'C2:C1000')  # Apply to Unit column
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
    
    # Save the template
    wb.save(filename)
    print(f"GEstimator Schedule template created: {filename}")
    return filename

def create_measurement_template(filename="gestimator_measurement_template.xlsx"):
    """Create a template for Measurements"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Measurements_Template"
    
    # Headers
    headers = ['Item_Code', 'Description', 'Nos', 'Length', 'Breadth', 'Height', 'Quantity', 'Unit', 'Remarks']
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add sample data
    sample_data = [
        ['1', 'Foundation excavation - Block A', 4, 20, 15, 1.5, '=C2*D2*E2*F2', 'Cum', 'Foundation trenches'],
        ['1', 'Foundation excavation - Block B', 2, 30, 10, 1.5, '=C3*D3*E3*F3', 'Cum', 'Foundation trenches'],
        ['2', 'Concrete footing - Block A', 8, 2, 2, 0.6, '=C4*D4*E4*F4', 'Cum', 'M25 grade'],
        ['2', 'Concrete footing - Block B', 6, 2.5, 2.5, 0.8, '=C5*D5*E5*F5', 'Cum', 'M25 grade'],
        ['3', 'Brick wall - Ground floor', 1, 120, 1, 3, '=C6*D6*E6*F6', 'Cum', 'External walls'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Set column widths
    column_widths = [12, 35, 8, 10, 10, 10, 12, 8, 25]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border
    
    # Save the template
    wb.save(filename)
    print(f"GEstimator Measurement template created: {filename}")
    return filename

def create_analysis_template(filename="gestimator_analysis_template.xlsx"):
    """Create a template for Rate Analysis"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis_Template"
    
    # Headers
    headers = ['Item_Code', 'Resource_Type', 'Description', 'Unit', 'Rate', 'Quantity', 'Amount']
    ws.append(headers)
    
    # Format headers
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add sample data
    sample_data = [
        ['1', 'MATERIAL', 'Cement', 'Bag', 450, 8.5, '=E2*F2'],
        ['1', 'MATERIAL', 'Sand', 'Cum', 800, 1.2, '=E3*F3'],
        ['1', 'MATERIAL', 'Aggregate', 'Cum', 1200, 2.4, '=E4*F4'],
        ['1', 'LABOUR', 'Mason', 'Day', 800, 2, '=E5*F5'],
        ['1', 'LABOUR', 'Helper', 'Day', 500, 4, '=E6*F6'],
        ['1', 'EQUIPMENT', 'Concrete mixer', 'Hour', 120, 8, '=E7*F7'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Set column widths
    column_widths = [12, 15, 35, 8, 12, 12, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Add data validation for Resource_Type column
    resource_validation = DataValidation(
        type="list",
        formula1='"MATERIAL,LABOUR,EQUIPMENT,OVERHEAD"',
        allow_blank=False
    )
    resource_validation.error = "Please select from the dropdown list"
    resource_validation.errorTitle = "Invalid Resource Type"
    ws.add_data_validation(resource_validation)
    resource_validation.add(f'B2:B1000')  # Apply to Resource_Type column
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
    
    # Save the template
    wb.save(filename)
    print(f"GEstimator Analysis template created: {filename}")
    return filename

def create_complete_template(filename="gestimator_complete_template.xlsx"):
    """Create a complete template with all sheets"""
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Schedule sheet
    ws_schedule = wb.create_sheet("Schedule")
    headers = ['Code', 'Description', 'Unit', 'Rate', 'Qty', 'Amount', 'Remarks']
    ws_schedule.append(headers)
    
    sample_schedule = [
        ['1', 'Earth work excavation', 'Cum', 178, 720, 128160, 'Foundation work'],
        ['2', 'Concrete M25', 'Cum', 5003, 34.8, 174104, 'RCC work'],
        ['3', 'Brick work', 'Sqm', 299, 1250, 373750, 'Masonry work'],
    ]
    
    for row_data in sample_schedule:
        ws_schedule.append(row_data)
    
    # Create Measurements sheet
    ws_measurements = wb.create_sheet("Measurements")
    headers = ['Item_Code', 'Description', 'Nos', 'Length', 'Breadth', 'Height', 'Quantity', 'Unit', 'Remarks']
    ws_measurements.append(headers)
    
    sample_measurements = [
        ['1', 'Foundation excavation', 4, 20, 15, 1.5, '=C2*D2*E2*F2', 'Cum', 'Block A'],
        ['2', 'Concrete footing', 8, 2, 2, 0.6, '=C3*D3*E3*F3', 'Cum', 'Footings'],
    ]
    
    for row_data in sample_measurements:
        ws_measurements.append(row_data)
    
    # Create Analysis sheet
    ws_analysis = wb.create_sheet("Analysis")
    headers = ['Item_Code', 'Resource_Type', 'Description', 'Unit', 'Rate', 'Quantity', 'Amount']
    ws_analysis.append(headers)
    
    sample_analysis = [
        ['1', 'MATERIAL', 'Cement', 'Bag', 450, 8.5, '=E2*F2'],
        ['1', 'LABOUR', 'Mason', 'Day', 800, 2, '=E3*F3'],
    ]
    
    for row_data in sample_analysis:
        ws_analysis.append(row_data)
    
    # Format all sheets
    for ws in wb.worksheets:
        # Header formatting
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        if ws.title == "Schedule":
            widths = [12, 50, 8, 12, 12, 15, 25]
        elif ws.title == "Measurements":
            widths = [12, 35, 8, 10, 10, 10, 12, 8, 25]
        else:  # Analysis
            widths = [12, 15, 35, 8, 12, 12, 15]
        
        for col, width in enumerate(widths, 1):
            if col <= ws.max_column:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Save the template
    wb.save(filename)
    print(f"Complete GEstimator template created: {filename}")
    return filename

def main():
    parser = argparse.ArgumentParser(description='Create GEstimator Excel templates')
    parser.add_argument('--type', choices=['schedule', 'measurement', 'analysis', 'complete'], 
                       default='complete', help='Type of template to create')
    parser.add_argument('-o', '--output', help='Output filename')
    
    args = parser.parse_args()
    
    if args.type == 'schedule':
        filename = args.output or 'gestimator_schedule_template.xlsx'
        create_schedule_template(filename)
    elif args.type == 'measurement':
        filename = args.output or 'gestimator_measurement_template.xlsx'
        create_measurement_template(filename)
    elif args.type == 'analysis':
        filename = args.output or 'gestimator_analysis_template.xlsx'
        create_analysis_template(filename)
    else:  # complete
        filename = args.output or 'gestimator_complete_template.xlsx'
        create_complete_template(filename)
    
    print(f"\nTemplate created successfully!")
    print(f"You can now:")
    print(f"1. Fill in your data following the sample format")
    print(f"2. Import the file into GEstimator")
    print(f"3. Use as a template for future projects")

if __name__ == "__main__":
    main()
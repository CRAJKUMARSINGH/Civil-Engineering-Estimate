#!/usr/bin/env python3
"""
Verify the converted GEstimator format file
"""
import openpyxl
import os

def verify_conversion():
    filename = 'geodesic_dome_gestimator.xlsx'
    print(f"Verifying converted GEstimator file: {filename}")
    print("=" * 50)
    
    # Check if file exists
    if not os.path.exists(filename):
        print(f"Error: File '{filename}' not found!")
        print("Current directory contents:")
        for file in os.listdir('.'):
            if file.endswith('.xlsx'):
                print(f"  - {file}")
        return
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        print(f"Excel file loaded successfully!")
        print(f"Sheet names: {wb.sheetnames}")
        
        # Check the Schedule sheet
        if 'Schedule' not in wb.sheetnames:
            print("Warning: 'Schedule' sheet not found!")
            return
            
        ws = wb['Schedule']
        print(f"Schedule sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Get column headers
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers.append(header)
        print(f"Headers: {headers}")
        
        # Show first 10 data rows
        print("\nFirst 10 data rows:")
        for row in range(2, min(12, ws.max_row + 1)):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row-1}: {row_data}")
        
        # Count total data rows
        total_rows = ws.max_row - 1  # Subtract header row
        print(f"\nTotal data rows: {total_rows}")
        
        print(f"\nVerification completed for {filename}")
        
    except Exception as e:
        print(f"Error reading converted file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify_conversion()
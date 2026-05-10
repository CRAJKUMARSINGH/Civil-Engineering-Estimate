#!/usr/bin/env python3
"""
Simple conversion script without GTK dependencies
"""

import sys
import os
from pathlib import Path
from decimal import Decimal

import openpyxl
import peewee
try:
    import appdirs
except ImportError:
    appdirs = None

# Constants
PROJECTS_DIR = "Attached_Assets/PROJECTS"
PROJECT_EXTENSION = ".eproj"
PROGRAM_NAME = "GEstimator"
PROGRAM_AUTHOR = "CPWD"
PROGRAM_VER = "1"
PROJECT_FILE_VER = "GESTIMATOR_FILE_REFERENCE_VER_2"

def get_orm_model(bind_database):
    """Simplified version of the ORM model"""
    
    class BaseModelSch(peewee.Model):
        class Meta:
            database = bind_database

    class ProjectTable(BaseModelSch):
        key = peewee.CharField(primary_key=True)
        value = peewee.CharField()

    class ScheduleCategoryTable(BaseModelSch):
        description = peewee.CharField(unique=True)
        order = peewee.IntegerField()

    class ResourceCategoryTable(BaseModelSch):
        description = peewee.CharField(unique=True)
        order = peewee.IntegerField()

    class ScheduleTable(BaseModelSch):
        code = peewee.CharField(unique=True)
        description = peewee.CharField()
        unit = peewee.CharField(null=True)
        rate = peewee.DecimalField(decimal_places=2, auto_round=True, null=True)
        qty = peewee.DecimalField(null=True)
        remarks = peewee.CharField(null=True)
        ana_remarks = peewee.CharField(null=True)
        category = peewee.ForeignKeyField(ScheduleCategoryTable, null=True, on_delete='CASCADE', backref='scheduleitems')
        parent = peewee.ForeignKeyField('self', on_delete='CASCADE', null=True, backref='children')
        order = peewee.IntegerField()
        suborder = peewee.IntegerField(null=True)
        colour = peewee.CharField(null=True)

    class ResourceTable(BaseModelSch):
        code = peewee.CharField(unique=True)
        description = peewee.CharField()
        unit = peewee.CharField()
        rate = peewee.DecimalField(decimal_places=2, auto_round=True)
        vat = peewee.DecimalField(decimal_places=2, auto_round=True, null=True)
        discount = peewee.DecimalField(decimal_places=2, auto_round=True, null=True)
        reference = peewee.CharField(null=True)
        category = peewee.ForeignKeyField(ResourceCategoryTable, null=True, on_delete='CASCADE', backref='resources')
        order = peewee.IntegerField()

    class SequenceTable(BaseModelSch):
        id_seq = peewee.IntegerField()
        id_sch = peewee.ForeignKeyField(ScheduleTable, on_delete='CASCADE', backref='sequences')
        itemtype = peewee.IntegerField()
        value = peewee.DecimalField(null=True)
        code = peewee.CharField(null=True)
        description = peewee.CharField(null=True)
        
        class MetaSequence:
            indexes = ((('id_seq', 'id_sch'), True),)

    class ResourceItemTable(BaseModelSch):
        id_sch = peewee.ForeignKeyField(ScheduleTable, on_delete='CASCADE', backref='resourceitems')
        id_seq = peewee.ForeignKeyField(SequenceTable, on_delete='CASCADE', backref='resourceitems')
        id_res = peewee.ForeignKeyField(ResourceTable, on_delete='CASCADE', backref='resourceitems')
        qty = peewee.DecimalField()
        remarks = peewee.CharField(null=True)

    return (BaseModelSch, ProjectTable, ScheduleCategoryTable, ResourceCategoryTable,
            ScheduleTable, ResourceTable, SequenceTable, ResourceItemTable)

def create_gestimator_database(output_path, project_name):
    """Create a new GEstimator database"""
    print(f"  Creating database: {output_path}")
    
    # Create database
    db = peewee.SqliteDatabase(str(output_path))
    
    # Get ORM models
    orm_models = get_orm_model(db)
    (BaseModelSch, ProjectTable, ScheduleCategoryTable, ResourceCategoryTable,
     ScheduleTable, ResourceTable, SequenceTable, ResourceItemTable) = orm_models
    
    # Connect and enable foreign keys
    db.connect()
    db.execute_sql('PRAGMA foreign_keys=ON;')
    
    # Create tables
    tables = [ProjectTable, ScheduleTable, ResourceTable,
             ScheduleCategoryTable, ResourceCategoryTable,
             SequenceTable, ResourceItemTable]
    db.create_tables(tables)
    
    # Set project settings
    ProjectTable.create(key='file_version', value=PROJECT_FILE_VER)
    ProjectTable.create(key='project_name', value=project_name)
    ProjectTable.create(key='project_item_code', value='')
    ProjectTable.create(key='project_resource_code', value='')
    ProjectTable.create(key='project_measurement', value='["Measurement", ["", []]]')
    
    db.commit()
    
    return db, orm_models

def parse_excel_project(excel_path):
    """Parse Excel project file and extract schedule data"""
    print(f"  Parsing Excel file: {excel_path.name}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Try to find the schedule sheet - look for sheets with "ABS" in the name
    schedule_sheet = None
    for sheet_name in wb.sheetnames:
        if 'ABS' in sheet_name.upper():
            schedule_sheet = wb[sheet_name]
            print(f"  Found schedule sheet: {sheet_name}")
            break
    
    # If no ABS sheet found, try other common names
    if schedule_sheet is None:
        for sheet_name in wb.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['schedule', 'boq', 'estimate', 'item']):
                schedule_sheet = wb[sheet_name]
                print(f"  Found schedule sheet: {sheet_name}")
                break
    
    # If still no sheet found, use the first sheet
    if schedule_sheet is None:
        schedule_sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
        if schedule_sheet is None:
            print("  Error: No worksheet found in Excel file")
            wb.close()
            return []
        print(f"  Using first sheet: {schedule_sheet.title}")
    
    # Parse schedule items
    items = []
    
    # Look for the header row in the schedule sheet
    header_row = None
    code_col = None
    desc_col = None
    qty_col = None
    unit_col = None
    rate_col = None
    amount_col = None
    
    # Scan first 30 rows to find the header
    for row_num in range(1, min(31, schedule_sheet.max_row + 1)):
        row_data = []
        for col_num in range(1, min(15, schedule_sheet.max_column + 1)):
            cell = schedule_sheet.cell(row=row_num, column=col_num)
            value = str(cell.value).strip() if cell.value is not None else ''
            row_data.append(value.lower())
        
        # Look for header patterns
        header_keywords = ['s.no', 's.n.', 'particular', 'quantity', 'qty', 'unit', 'rate', 'amount']
        matches = sum(1 for v in row_data if any(keyword in v for keyword in header_keywords))
        
        if matches >= 3:  # At least 3 header keywords found
            header_row = row_num
            print(f"  Found header at row {header_row}: {row_data}")
            
            # Identify columns
            for idx, header in enumerate(row_data):
                if code_col is None and ('s.no' in header or 's.n.' in header):
                    code_col = idx + 1  # 1-based indexing
                elif desc_col is None and 'particular' in header:
                    desc_col = idx + 1
                elif qty_col is None and ('quantity' in header or 'qty' in header):
                    qty_col = idx + 1
                elif unit_col is None and 'unit' in header:
                    unit_col = idx + 1
                elif rate_col is None and 'rate' in header:
                    rate_col = idx + 1
                elif amount_col is None and 'amount' in header:
                    amount_col = idx + 1
            
            print(f"  Column mapping - Code:{code_col}, Desc:{desc_col}, Qty:{qty_col}, Unit:{unit_col}, Rate:{rate_col}, Amount:{amount_col}")
            break
    
    # If no header found, try to guess based on common patterns
    if header_row is None:
        print("  Warning: Could not find header row, trying to guess data structure")
        # Assume row 5 is the header based on our analysis
        header_row = 5
        code_col = 1
        desc_col = 2
        qty_col = 3
        unit_col = 6
        rate_col = 5
        amount_col = 7
        print(f"  Using assumed column mapping - Code:{code_col}, Desc:{desc_col}, Qty:{qty_col}, Unit:{unit_col}, Rate:{rate_col}, Amount:{amount_col}")
    
    # Parse data rows starting after the header
    start_row = header_row + 1
    print(f"  Parsing data starting from row {start_row}")
    
    for row_num in range(start_row, schedule_sheet.max_row + 1):
        try:
            # Get values from the identified columns
            code_cell = schedule_sheet.cell(row=row_num, column=code_col) if code_col else None
            desc_cell = schedule_sheet.cell(row=row_num, column=desc_col) if desc_col else None
            qty_cell = schedule_sheet.cell(row=row_num, column=qty_col) if qty_col else None
            unit_cell = schedule_sheet.cell(row=row_num, column=unit_col) if unit_col else None
            rate_cell = schedule_sheet.cell(row=row_num, column=rate_col) if rate_col else None
            
            # Get actual values
            code = code_cell.value if code_cell and code_cell.value is not None else None
            description = desc_cell.value if desc_cell and desc_cell.value is not None else None
            qty = qty_cell.value if qty_cell and qty_cell.value is not None else None
            unit = unit_cell.value if unit_cell and unit_cell.value is not None else None
            rate = rate_cell.value if rate_cell and rate_cell.value is not None else None
            
            # Skip empty rows
            if not code and not description:
                continue
                
            # Validate and convert
            if not code:
                code = f"ITEM_{row_num}"
            if not description:
                continue
            if not unit:
                unit = "Unit"
            
            code = str(code).strip()
            description = str(description).strip()[:1000]  # Limit length
            unit = str(unit).strip()
            
            # Convert rate and qty to Decimal
            try:
                rate = Decimal(str(rate)) if rate else Decimal('0')
            except:
                rate = Decimal('0')
            
            try:
                qty = Decimal(str(qty)) if qty else Decimal('0')
            except:
                qty = Decimal('0')
            
            # Skip rows that look like totals or headers
            if any(keyword in code.lower() for keyword in ['total', 'subtotal', 'grand total', 'say']):
                continue
                
            items.append({
                'code': code,
                'description': description,
                'unit': unit,
                'rate': rate,
                'qty': qty
            })
            
            # Debug: print first few items
            if len(items) <= 3:
                print(f"    Parsed item: {code} - {description[:50]}...")
            
        except Exception as e:
            print(f"  Warning: Error parsing row {row_num}: {e}")
            continue
    
    wb.close()
    print(f"  Extracted {len(items)} schedule items")
    
    return items

def insert_schedule_items(db, orm_models, items):
    """Insert schedule items into database"""
    (BaseModelSch, ProjectTable, ScheduleCategoryTable, ResourceCategoryTable,
     ScheduleTable, ResourceTable, SequenceTable, ResourceItemTable) = orm_models
    
    print(f"  Inserting {len(items)} items into database...")
    
    inserted = 0
    for idx, item in enumerate(items, 1):
        try:
            ScheduleTable.create(
                code=item['code'],
                description=item['description'],
                unit=item['unit'],
                rate=item['rate'],
                qty=item['qty'],
                remarks='',
                ana_remarks='',
                category=None,
                parent=None,
                order=idx,
                suborder=None,
                colour=None
            )
            inserted += 1
        except Exception as e:
            print(f"    Warning: Could not insert item {item['code']}: {e}")
    
    db.commit()
    print(f"  Successfully inserted {inserted} items")
    return inserted

def convert_excel_to_eproj(excel_path, output_dir):
    """Convert Excel project file to .eproj format"""
    print(f"\nConverting: {excel_path.name}")
    print("="*80)
    
    try:
        # Create output filename
        project_name = excel_path.stem
        output_filename = project_name + PROJECT_EXTENSION
        output_path = output_dir / output_filename
        
        # Check if already exists
        if output_path.exists():
            print(f"  Skipping: {output_filename} already exists")
            return False
        
        # Parse Excel file
        items = parse_excel_project(excel_path)
        
        if not items:
            print(f"  Error: No schedule items found in Excel file")
            return False
        
        # Create GEstimator database
        db, orm_models = create_gestimator_database(output_path, project_name)
        
        # Insert schedule items
        inserted = insert_schedule_items(db, orm_models, items)
        
        # Close database
        db.close()
        
        print(f"  ✓ Successfully created: {output_filename}")
        print(f"  Location: {output_path}")
        print(f"  Items: {inserted}")
        
        return True
        
    except Exception as e:
        print(f"  ✗ Error converting {excel_path.name}: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main conversion process"""
    print("="*80)
    print("Excel to GEstimator Project Converter")
    print("="*80)
    
    # Get target directory
    if appdirs is not None:
        dirs = appdirs.AppDirs(PROGRAM_NAME, PROGRAM_AUTHOR, version=PROGRAM_VER)
        user_data_dir = Path(dirs.user_data_dir)
    else:
        # Fallback to local directory if appdirs is not available
        user_data_dir = Path.home() / f".{PROGRAM_NAME.lower()}"
    
    target_dir = user_data_dir / 'projects'
    target_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\nTarget directory: {target_dir}")
    
    # Find Excel project files
    projects_dir = Path(PROJECTS_DIR)
    excel_files = list(projects_dir.glob("*.xlsx"))
    
    print(f"\nFound {len(excel_files)} Excel project files:")
    for f in excel_files:
        print(f"  - {f.name}")
    
    # Convert each file
    print(f"\n{'='*80}")
    print("Starting conversion...")
    print(f"{'='*80}")
    
    converted = 0
    skipped = 0
    failed = 0
    
    for excel_file in excel_files:
        result = convert_excel_to_eproj(excel_file, target_dir)
        if result:
            converted += 1
        elif result is False:
            skipped += 1
        else:
            failed += 1
    
    # Summary
    print(f"\n{'='*80}")
    print("CONVERSION SUMMARY")
    print(f"{'='*80}")
    print(f"Total files: {len(excel_files)}")
    print(f"Converted: {converted}")
    print(f"Skipped: {skipped}")
    print(f"Failed: {failed}")
    print(f"\nConverted projects saved to: {target_dir}")
    print("="*80)
    
    return 0 if failed == 0 else 1

if __name__ == '__main__':
    try:
        sys.exit(main())
    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
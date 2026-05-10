#!/usr/bin/env python3
"""
Detailed analysis of Excel files to understand data structure
"""

import openpyxl
from pathlib import Path

def analyze_excel_file(file_path):
    """Analyze the structure of an Excel file in detail"""
    print(f"\nAnalyzing: {file_path.name}")
    print("="*60)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        print(f"Number of sheets: {len(wb.sheetnames)}")
        print(f"Sheet names: {wb.sheetnames}")
        
        # Analyze each sheet in detail
        for sheet_name in wb.sheetnames:
            print(f"\nSheet: {sheet_name}")
            sheet = wb[sheet_name]
            
            print(f"  Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Show more rows to find the data
            print("  First 20 rows:")
            for row_num in range(1, min(21, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = str(cell.value)[:30] if cell.value is not None else ''
                    row_data.append(value)
                print(f"    Row {row_num:2d}: {row_data}")
                
            # Try to find where the actual data starts
            print("\n  Looking for data patterns...")
            for row_num in range(1, min(31, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    row_data.append(value)
                
                # Check if this row might contain header or data
                non_empty_count = sum(1 for v in row_data if v is not None and str(v).strip())
                if non_empty_count >= 4:  # At least 4 non-empty cells
                    print(f"    Row {row_num:2d} (non-empty: {non_empty_count}): {[str(v)[:20] if v is not None else '' for v in row_data]}")
                    
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing {file_path.name}: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main analysis function"""
    projects_dir = Path("Attached_Assets/PROJECTS")
    excel_files = list(projects_dir.glob("*.xlsx"))
    
    print(f"Found {len(excel_files)} Excel files:")
    for f in excel_files:
        print(f"  - {f.name}")
    
    print("\n" + "="*80)
    print("DETAILED ANALYSIS OF EXCEL FILES")
    print("="*80)
    
    for excel_file in excel_files:
        analyze_excel_file(excel_file)

if __name__ == '__main__':
    main()
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Enhanced Excel Exporter with customizable templates and formatting
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .. import misc

log = logging.getLogger(__name__)

class EnhancedExcelExporter:
    """Enhanced Excel exporter with template support and advanced formatting"""
    
    def __init__(self, template_manager=None):
        """Initialize exporter with optional template manager"""
        self.template_manager = template_manager
        
        # Default styling
        self.styles = {
            'header': {
                'font': Font(bold=True, size=12, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'subheader': {
                'font': Font(bold=True, size=11),
                'fill': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'data': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'number': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                'number_format': '#,##0.00'
            },
            'total': {
                'font': Font(bold=True, size=10),
                'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium')
                ),
                'number_format': '#,##0.00'
            }
        }
    
    def export_with_template(self, estimate_data: Dict, template_name: str, 
                           output_path: Path, organization_info: Dict = None) -> bool:
        """
        Export estimate using predefined template
        
        Args:
            estimate_data: Dictionary containing estimate data
            template_name: Name of template to use
            output_path: Path for output Excel file
            organization_info: Optional organization branding info
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Load template if template manager is available
            template_config = None
            if self.template_manager:
                template_config = self.template_manager.load_template(template_name)
            
            if not template_config:
                # Use default template
                template_config = self._get_default_template_config()
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create sheets based on template configuration
            for sheet_config in template_config.get('sheets', []):
                self._create_sheet_from_config(workbook, sheet_config, estimate_data)
            
            # Apply organization branding if provided
            if organization_info:
                self._apply_organization_branding(workbook, organization_info)
            
            # Save workbook
            workbook.save(str(output_path))
            
            log.info(f"Estimate exported using template '{template_name}' to: {output_path}")
            return True
            
        except Exception as e:
            log.error(f"Error exporting with template: {e}")
            return False
    
    def export_schedule_enhanced(self, schedule_data: List[Dict], output_path: Path,
                               include_analysis: bool = True, 
                               include_measurements: bool = True,
                               formatting_options: Dict = None) -> bool:
        """
        Export schedule with enhanced formatting and options
        
        Args:
            schedule_data: List of schedule item dictionaries
            output_path: Path for output Excel file
            include_analysis: Whether to include analysis sheets
            include_measurements: Whether to include measurement sheets
            formatting_options: Custom formatting options
            
        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = openpyxl.Workbook()
            
            # Apply custom formatting if provided
            if formatting_options:
                self._update_styles(formatting_options)
            
            # Create main schedule sheet
            self._create_schedule_sheet(workbook, schedule_data)
            
            # Create summary sheet
            self._create_summary_sheet(workbook, schedule_data)
            
            # Create analysis sheets if requested
            if include_analysis:
                self._create_analysis_sheets(workbook, schedule_data)
            
            # Create measurement sheets if requested
            if include_measurements:
                self._create_measurement_sheets(workbook, schedule_data)
            
            # Remove default sheet if it exists
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Set active sheet to schedule
            if 'Schedule' in workbook.sheetnames:
                workbook.active = workbook['Schedule']
            
            # Save workbook
            workbook.save(str(output_path))
            
            log.info(f"Enhanced schedule exported to: {output_path}")
            return True
            
        except Exception as e:
            log.error(f"Error exporting enhanced schedule: {e}")
            return False
    
    def _create_schedule_sheet(self, workbook: openpyxl.Workbook, 
                             schedule_data: List[Dict]):
        """Create main schedule sheet"""
        worksheet = workbook.create_sheet("Schedule")
        
        # Headers
        headers = [
            'S.No.', 'Item Code', 'Description', 'Unit', 'Quantity', 
            'Rate', 'Amount', 'Category', 'Remarks'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            self._apply_style(cell, 'header')
        
        # Write data
        total_amount = 0
        for row, item in enumerate(schedule_data, 2):
            # Serial number
            worksheet.cell(row=row, column=1, value=row-1)
            
            # Item data
            worksheet.cell(row=row, column=2, value=item.get('code', ''))
            worksheet.cell(row=row, column=3, value=item.get('description', ''))
            worksheet.cell(row=row, column=4, value=item.get('unit', ''))
            
            qty = float(item.get('qty', 0))
            rate = float(item.get('rate', 0))
            amount = qty * rate
            total_amount += amount
            
            worksheet.cell(row=row, column=5, value=qty)
            worksheet.cell(row=row, column=6, value=rate)
            worksheet.cell(row=row, column=7, value=amount)
            worksheet.cell(row=row, column=8, value=item.get('category', ''))
            worksheet.cell(row=row, column=9, value=item.get('remarks', ''))
            
            # Apply data styling
            for col in range(1, 10):
                cell = worksheet.cell(row=row, column=col)
                if col in [5, 6, 7]:  # Numeric columns
                    self._apply_style(cell, 'number')
                else:
                    self._apply_style(cell, 'data')
        
        # Add total row
        total_row = len(schedule_data) + 2
        worksheet.cell(row=total_row, column=6, value="TOTAL:")
        worksheet.cell(row=total_row, column=7, value=total_amount)
        
        # Style total row
        for col in [6, 7]:
            cell = worksheet.cell(row=total_row, column=col)
            self._apply_style(cell, 'total')
        
        # Set column widths
        column_widths = [8, 15, 50, 10, 12, 15, 15, 15, 25]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, 
                            schedule_data: List[Dict]):
        """Create summary sheet with statistics"""
        worksheet = workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_items = len(schedule_data)
        total_amount = sum(float(item.get('qty', 0)) * float(item.get('rate', 0)) 
                          for item in schedule_data)
        
        categories = {}
        for item in schedule_data:
            category = item.get('category', 'Uncategorized')
            if category not in categories:
                categories[category] = {'count': 0, 'amount': 0}
            categories[category]['count'] += 1
            categories[category]['amount'] += float(item.get('qty', 0)) * float(item.get('rate', 0))
        
        # Project information
        row = 1
        worksheet.cell(row=row, column=1, value="PROJECT SUMMARY")
        self._apply_style(worksheet.cell(row=row, column=1), 'header')
        worksheet.merge_cells(f'A{row}:D{row}')
        
        row += 2
        summary_data = [
            ['Generated On:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Items:', total_items],
            ['Total Amount:', total_amount],
            ['Categories:', len(categories)]
        ]
        
        for label, value in summary_data:
            worksheet.cell(row=row, column=1, value=label)
            worksheet.cell(row=row, column=2, value=value)
            
            self._apply_style(worksheet.cell(row=row, column=1), 'subheader')
            if isinstance(value, (int, float)):
                self._apply_style(worksheet.cell(row=row, column=2), 'number')
            else:
                self._apply_style(worksheet.cell(row=row, column=2), 'data')
            
            row += 1
        
        # Category breakdown
        row += 2
        worksheet.cell(row=row, column=1, value="CATEGORY BREAKDOWN")
        self._apply_style(worksheet.cell(row=row, column=1), 'header')
        worksheet.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Category', 'Items', 'Amount', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            self._apply_style(cell, 'subheader')
        
        row += 1
        for category, data in sorted(categories.items()):
            percentage = (data['amount'] / total_amount * 100) if total_amount > 0 else 0
            
            worksheet.cell(row=row, column=1, value=category)
            worksheet.cell(row=row, column=2, value=data['count'])
            worksheet.cell(row=row, column=3, value=data['amount'])
            worksheet.cell(row=row, column=4, value=f"{percentage:.1f}%")
            
            # Apply styling
            self._apply_style(worksheet.cell(row=row, column=1), 'data')
            self._apply_style(worksheet.cell(row=row, column=2), 'number')
            self._apply_style(worksheet.cell(row=row, column=3), 'number')
            self._apply_style(worksheet.cell(row=row, column=4), 'data')
            
            row += 1
        
        # Set column widths
        column_widths = [20, 15, 20, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width
    
    def _create_analysis_sheets(self, workbook: openpyxl.Workbook, 
                              schedule_data: List[Dict]):
        """Create analysis sheets for items with analysis data"""
        # This would integrate with the existing analysis functionality
        # For now, create a placeholder sheet
        worksheet = workbook.create_sheet("Analysis")
        
        worksheet.cell(row=1, column=1, value="Analysis of Rates")
        self._apply_style(worksheet.cell(row=1, column=1), 'header')
        
        worksheet.cell(row=3, column=1, value="Analysis sheets will be generated for items with rate analysis data.")
        self._apply_style(worksheet.cell(row=3, column=1), 'data')
    
    def _create_measurement_sheets(self, workbook: openpyxl.Workbook, 
                                 schedule_data: List[Dict]):
        """Create measurement sheets"""
        # This would integrate with the existing measurement functionality
        # For now, create a placeholder sheet
        worksheet = workbook.create_sheet("Measurements")
        
        worksheet.cell(row=1, column=1, value="Details of Measurements")
        self._apply_style(worksheet.cell(row=1, column=1), 'header')
        
        worksheet.cell(row=3, column=1, value="Measurement details will be generated for items with measurement data.")
        self._apply_style(worksheet.cell(row=3, column=1), 'data')
    
    def _create_sheet_from_config(self, workbook: openpyxl.Workbook, 
                                sheet_config: Dict, estimate_data: Dict):
        """Create sheet from template configuration"""
        sheet_name = sheet_config.get('name', 'Sheet')
        worksheet = workbook.create_sheet(sheet_name)
        
        # Apply sheet-specific configuration
        columns = sheet_config.get('columns', [])
        formatting = sheet_config.get('formatting', {})
        
        # This would be expanded based on template requirements
        # For now, create a basic sheet
        if sheet_name.lower() == 'schedule':
            self._create_schedule_sheet(workbook, estimate_data.get('schedule', []))
    
    def _apply_organization_branding(self, workbook: openpyxl.Workbook, 
                                   organization_info: Dict):
        """Apply organization branding to workbook"""
        # Add organization header to each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Insert rows at top for organization info
            worksheet.insert_rows(1, 3)
            
            # Organization name
            org_name = organization_info.get('name', 'Organization Name')
            worksheet.cell(row=1, column=1, value=org_name)
            worksheet.merge_cells('A1:I1')
            
            # Style organization header
            cell = worksheet.cell(row=1, column=1)
            cell.font = Font(bold=True, size=16, color='000080')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Project info
            project_info = organization_info.get('project_info', '')
            if project_info:
                worksheet.cell(row=2, column=1, value=project_info)
                worksheet.merge_cells('A2:I2')
                
                cell = worksheet.cell(row=2, column=1)
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_style(self, cell, style_name: str):
        """Apply predefined style to cell"""
        if style_name in self.styles:
            style = self.styles[style_name]
            
            if 'font' in style:
                cell.font = style['font']
            if 'fill' in style:
                cell.fill = style['fill']
            if 'alignment' in style:
                cell.alignment = style['alignment']
            if 'border' in style:
                cell.border = style['border']
            if 'number_format' in style:
                cell.number_format = style['number_format']
    
    def _update_styles(self, formatting_options: Dict):
        """Update styles with custom formatting options"""
        for style_name, style_updates in formatting_options.items():
            if style_name in self.styles:
                self.styles[style_name].update(style_updates)
    
    def _get_default_template_config(self) -> Dict:
        """Get default template configuration"""
        return {
            'name': 'Default Template',
            'sheets': [
                {
                    'name': 'Schedule',
                    'columns': [
                        {'name': 'Code', 'width': 15},
                        {'name': 'Description', 'width': 50},
                        {'name': 'Unit', 'width': 10},
                        {'name': 'Quantity', 'width': 12},
                        {'name': 'Rate', 'width': 15},
                        {'name': 'Amount', 'width': 15}
                    ],
                    'formatting': {
                        'header_color': '366092',
                        'alternate_rows': True
                    }
                }
            ]
        }
    
    def create_comparison_report(self, comparison_data: Dict, 
                               output_path: Path) -> bool:
        """
        Create Excel comparison report from batch processing results
        
        Args:
            comparison_data: Comparison matrix data
            output_path: Path for output Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = openpyxl.Workbook()
            
            # Overview sheet
            self._create_comparison_overview(workbook, comparison_data)
            
            # Common items sheet
            if comparison_data.get('common_items'):
                self._create_common_items_sheet(workbook, comparison_data)
            
            # Unique items sheet
            if comparison_data.get('unique_items'):
                self._create_unique_items_sheet(workbook, comparison_data)
            
            # Rate differences sheet
            if comparison_data.get('rate_differences'):
                self._create_rate_differences_sheet(workbook, comparison_data)
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            # Save workbook
            workbook.save(str(output_path))
            
            log.info(f"Comparison report exported to: {output_path}")
            return True
            
        except Exception as e:
            log.error(f"Error creating comparison report: {e}")
            return False
    
    def _create_comparison_overview(self, workbook: openpyxl.Workbook, 
                                  comparison_data: Dict):
        """Create comparison overview sheet"""
        worksheet = workbook.create_sheet("Overview")
        
        # Title
        worksheet.cell(row=1, column=1, value="File Comparison Report")
        self._apply_style(worksheet.cell(row=1, column=1), 'header')
        worksheet.merge_cells('A1:D1')
        
        # Summary statistics
        row = 3
        stats = [
            ['Files Compared:', len(comparison_data.get('files', []))],
            ['Total Unique Items:', comparison_data.get('total_items', 0)],
            ['Common Items:', len(comparison_data.get('common_items', []))],
            ['Rate Differences Found:', len(comparison_data.get('rate_differences', []))]
        ]
        
        for label, value in stats:
            worksheet.cell(row=row, column=1, value=label)
            worksheet.cell(row=row, column=2, value=value)
            
            self._apply_style(worksheet.cell(row=row, column=1), 'subheader')
            self._apply_style(worksheet.cell(row=row, column=2), 'data')
            
            row += 1
        
        # Files list
        row += 2
        worksheet.cell(row=row, column=1, value="Files Included:")
        self._apply_style(worksheet.cell(row=row, column=1), 'subheader')
        
        row += 1
        for filename in comparison_data.get('files', []):
            worksheet.cell(row=row, column=1, value=filename)
            self._apply_style(worksheet.cell(row=row, column=1), 'data')
            row += 1
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 25
        worksheet.column_dimensions['B'].width = 20
    
    def _create_common_items_sheet(self, workbook: openpyxl.Workbook, 
                                 comparison_data: Dict):
        """Create common items sheet"""
        worksheet = workbook.create_sheet("Common Items")
        
        # Headers
        worksheet.cell(row=1, column=1, value="Item Code")
        self._apply_style(worksheet.cell(row=1, column=1), 'header')
        
        # Write common items
        for row, code in enumerate(comparison_data['common_items'], 2):
            worksheet.cell(row=row, column=1, value=code)
            self._apply_style(worksheet.cell(row=row, column=1), 'data')
        
        worksheet.column_dimensions['A'].width = 20
    
    def _create_unique_items_sheet(self, workbook: openpyxl.Workbook, 
                                 comparison_data: Dict):
        """Create unique items sheet"""
        worksheet = workbook.create_sheet("Unique Items")
        
        # Headers
        worksheet.cell(row=1, column=1, value="File Name")
        worksheet.cell(row=1, column=2, value="Unique Item Code")
        
        for col in [1, 2]:
            self._apply_style(worksheet.cell(row=1, column=col), 'header')
        
        # Write unique items
        row = 2
        for filename, unique_codes in comparison_data['unique_items'].items():
            for code in unique_codes:
                worksheet.cell(row=row, column=1, value=filename)
                worksheet.cell(row=row, column=2, value=code)
                
                for col in [1, 2]:
                    self._apply_style(worksheet.cell(row=row, column=col), 'data')
                
                row += 1
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 20
    
    def _create_rate_differences_sheet(self, workbook: openpyxl.Workbook, 
                                     comparison_data: Dict):
        """Create rate differences sheet"""
        worksheet = workbook.create_sheet("Rate Differences")
        
        # Headers
        headers = ['Item Code', 'Min Rate', 'Max Rate', 'Difference %']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
            self._apply_style(worksheet.cell(row=1, column=col), 'header')
        
        # Write rate differences
        for row, diff_data in enumerate(comparison_data['rate_differences'], 2):
            worksheet.cell(row=row, column=1, value=diff_data['code'])
            worksheet.cell(row=row, column=2, value=diff_data['min_rate'])
            worksheet.cell(row=row, column=3, value=diff_data['max_rate'])
            worksheet.cell(row=row, column=4, value=f"{diff_data['difference_percent']:.1f}%")
            
            # Apply styling
            self._apply_style(worksheet.cell(row=row, column=1), 'data')
            for col in [2, 3]:
                self._apply_style(worksheet.cell(row=row, column=col), 'number')
            self._apply_style(worksheet.cell(row=row, column=4), 'data')
        
        # Set column widths
        column_widths = [20, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(col)].width = width
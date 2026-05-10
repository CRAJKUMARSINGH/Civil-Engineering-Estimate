#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Enhanced Excel Importer with partial row selection and preview functionality
"""

import logging
from pathlib import Path
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Tuple, Optional, Any
import openpyxl
from dataclasses import dataclass

from .. import misc
from ..data import schedule

log = logging.getLogger(__name__)

@dataclass
class ImportPreviewItem:
    """Data class for import preview items"""
    row_number: int
    code: str
    description: str
    unit: str
    rate: Decimal
    qty: Decimal
    category: Optional[str] = None
    remarks: Optional[str] = None
    selected: bool = True
    validation_errors: List[str] = None
    
    def __post_init__(self):
        if self.validation_errors is None:
            self.validation_errors = []

class EnhancedExcelImporter:
    """Enhanced Excel importer with preview and partial selection capabilities"""
    
    def __init__(self, ssr_manager=None, template_manager=None):
        """Initialize importer with optional SSR and template managers"""
        self.ssr_manager = ssr_manager
        self.template_manager = template_manager
        self.required_columns = ['Code', 'Description', 'Unit', 'Rate', 'Quantity']
        self.optional_columns = ['Category', 'Remarks']
        
    def analyze_excel_file(self, filepath: Path) -> Dict[str, Any]:
        """
        Analyze Excel file structure and return metadata
        
        Returns:
            Dict containing file analysis results
        """
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            
            analysis = {
                'filename': filepath.name,
                'sheets': [],
                'total_sheets': len(workbook.sheetnames),
                'recommended_sheet': None,
                'file_size': filepath.stat().st_size,
                'success': True,
                'errors': []
            }
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_info = self._analyze_worksheet(worksheet, sheet_name)
                analysis['sheets'].append(sheet_info)
                
                # Determine recommended sheet (prefer sheets with schedule-like content)
                if not analysis['recommended_sheet']:
                    if any(keyword in sheet_name.lower() for keyword in 
                          ['schedule', 'boq', 'estimate', 'item', 'abs']):
                        analysis['recommended_sheet'] = sheet_name
            
            # If no recommended sheet found, use first sheet with data
            if not analysis['recommended_sheet'] and analysis['sheets']:
                for sheet_info in analysis['sheets']:
                    if sheet_info['data_rows'] > 0:
                        analysis['recommended_sheet'] = sheet_info['name']
                        break
            
            workbook.close()
            return analysis
            
        except Exception as e:
            log.error(f"Error analyzing Excel file {filepath}: {e}")
            return {
                'filename': filepath.name,
                'success': False,
                'errors': [str(e)]
            }
    
    def _analyze_worksheet(self, worksheet, sheet_name: str) -> Dict[str, Any]:
        """Analyze individual worksheet structure"""
        sheet_info = {
            'name': sheet_name,
            'max_row': worksheet.max_row,
            'max_column': worksheet.max_column,
            'header_row': None,
            'data_rows': 0,
            'column_mapping': {},
            'sample_data': []
        }
        
        # Find header row
        header_row = self._find_header_row(worksheet)
        if header_row:
            sheet_info['header_row'] = header_row
            sheet_info['column_mapping'] = self._map_columns(worksheet, header_row)
            sheet_info['data_rows'] = max(0, worksheet.max_row - header_row)
            
            # Get sample data (first 5 rows after header)
            sample_rows = min(5, sheet_info['data_rows'])
            for row_num in range(header_row + 1, header_row + 1 + sample_rows):
                row_data = []
                for col_num in range(1, min(10, worksheet.max_column + 1)):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    value = str(cell.value) if cell.value is not None else ''
                    row_data.append(value[:50])  # Truncate for preview
                sheet_info['sample_data'].append(row_data)
        
        return sheet_info
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        """Find the header row in worksheet"""
        header_keywords = ['code', 'description', 'particular', 'quantity', 'qty', 'unit', 'rate', 'amount']
        
        # Scan first 30 rows
        for row_num in range(1, min(31, worksheet.max_row + 1)):
            row_data = []
            for col_num in range(1, min(15, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = str(cell.value).strip().lower() if cell.value is not None else ''
                row_data.append(value)
            
            # Check if this row contains header keywords
            matches = sum(1 for value in row_data 
                         if any(keyword in value for keyword in header_keywords))
            
            if matches >= 3:  # At least 3 header keywords found
                return row_num
        
        return None
    
    def _map_columns(self, worksheet, header_row: int) -> Dict[str, int]:
        """Map column names to indices"""
        column_mapping = {}
        
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            header = str(cell.value).strip().lower() if cell.value is not None else ''
            
            # Map common column variations
            if any(keyword in header for keyword in ['code', 's.no', 's.n.']):
                column_mapping['Code'] = col_num
            elif any(keyword in header for keyword in ['description', 'particular', 'item']):
                column_mapping['Description'] = col_num
            elif 'unit' in header:
                column_mapping['Unit'] = col_num
            elif any(keyword in header for keyword in ['rate', 'unit rate']):
                column_mapping['Rate'] = col_num
            elif any(keyword in header for keyword in ['quantity', 'qty']):
                column_mapping['Quantity'] = col_num
            elif 'category' in header:
                column_mapping['Category'] = col_num
            elif any(keyword in header for keyword in ['remarks', 'note']):
                column_mapping['Remarks'] = col_num
        
        return column_mapping
    
    def preview_import(self, filepath: Path, sheet_name: str = None, 
                      start_row: int = None, end_row: int = None,
                      selected_rows: List[int] = None) -> List[ImportPreviewItem]:
        """
        Preview Excel import with validation
        
        Args:
            filepath: Path to Excel file
            sheet_name: Name of sheet to import (None for active sheet)
            start_row: Start row for range import
            end_row: End row for range import  
            selected_rows: List of specific row numbers to import
            
        Returns:
            List of ImportPreviewItem objects
        """
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # Find header row and column mapping
            header_row = self._find_header_row(worksheet)
            if not header_row:
                raise ValueError("Could not find header row in worksheet")
            
            column_mapping = self._map_columns(worksheet, header_row)
            
            # Validate required columns
            missing_columns = [col for col in self.required_columns 
                             if col not in column_mapping]
            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
            
            # Determine rows to process
            if selected_rows:
                rows_to_process = selected_rows
            elif start_row and end_row:
                rows_to_process = list(range(start_row, end_row + 1))
            else:
                rows_to_process = list(range(header_row + 1, worksheet.max_row + 1))
            
            # Parse rows into preview items
            preview_items = []
            for row_num in rows_to_process:
                if row_num <= header_row:
                    continue
                    
                try:
                    item = self._parse_row_to_preview(worksheet, row_num, column_mapping)
                    if item:  # Skip empty rows
                        preview_items.append(item)
                except Exception as e:
                    log.warning(f"Error parsing row {row_num}: {e}")
            
            workbook.close()
            
            # Validate items and check against SSR if available
            if self.ssr_manager:
                self._validate_against_ssr(preview_items)
            
            return preview_items
            
        except Exception as e:
            log.error(f"Error previewing Excel import: {e}")
            raise
    
    def _parse_row_to_preview(self, worksheet, row_num: int, 
                            column_mapping: Dict[str, int]) -> Optional[ImportPreviewItem]:
        """Parse worksheet row to ImportPreviewItem"""
        
        # Extract values
        code = self._get_cell_value(worksheet, row_num, column_mapping.get('Code'))
        description = self._get_cell_value(worksheet, row_num, column_mapping.get('Description'))
        unit = self._get_cell_value(worksheet, row_num, column_mapping.get('Unit'))
        rate = self._get_cell_value(worksheet, row_num, column_mapping.get('Rate'))
        qty = self._get_cell_value(worksheet, row_num, column_mapping.get('Quantity'))
        category = self._get_cell_value(worksheet, row_num, column_mapping.get('Category'))
        remarks = self._get_cell_value(worksheet, row_num, column_mapping.get('Remarks'))
        
        # Skip empty rows
        if not code and not description:
            return None
        
        # Validate and convert data
        validation_errors = []
        
        # Code validation
        if not code:
            code = f"ITEM_{row_num}"
            validation_errors.append("Missing item code - auto-generated")
        else:
            code = str(code).strip()
        
        # Description validation
        if not description:
            validation_errors.append("Missing description")
            return None
        else:
            description = str(description).strip()
            if len(description) > misc.MAX_DESC_LEN:
                description = description[:misc.MAX_DESC_LEN]
                validation_errors.append(f"Description truncated to {misc.MAX_DESC_LEN} characters")
        
        # Unit validation
        if not unit:
            unit = "Unit"
            validation_errors.append("Missing unit - using default 'Unit'")
        else:
            unit = str(unit).strip()
        
        # Rate conversion
        try:
            rate = Decimal(str(rate)) if rate else Decimal('0')
        except (InvalidOperation, ValueError):
            rate = Decimal('0')
            validation_errors.append("Invalid rate value - using 0")
        
        # Quantity conversion
        try:
            qty = Decimal(str(qty)) if qty else Decimal('0')
        except (InvalidOperation, ValueError):
            qty = Decimal('0')
            validation_errors.append("Invalid quantity value - using 0")
        
        # Category and remarks
        category = str(category).strip() if category else None
        remarks = str(remarks).strip() if remarks else None
        
        return ImportPreviewItem(
            row_number=row_num,
            code=code,
            description=description,
            unit=unit,
            rate=rate,
            qty=qty,
            category=category,
            remarks=remarks,
            validation_errors=validation_errors
        )
    
    def _get_cell_value(self, worksheet, row: int, col: Optional[int]) -> Any:
        """Safely get cell value"""
        if col is None:
            return None
        try:
            return worksheet.cell(row=row, column=col).value
        except:
            return None
    
    def _validate_against_ssr(self, preview_items: List[ImportPreviewItem]):
        """Validate preview items against SSR database"""
        if not self.ssr_manager:
            return
        
        for item in preview_items:
            # Check if item exists in SSR
            ssr_match = self.ssr_manager.find_item_by_code(item.code)
            if not ssr_match:
                # Try fuzzy matching by description
                ssr_match = self.ssr_manager.fuzzy_match_description(
                    item.description, threshold=0.85
                )
                if ssr_match:
                    item.validation_errors.append(
                        f"Fuzzy matched to SSR item: {ssr_match['code']}"
                    )
                else:
                    item.validation_errors.append("Item not found in SSR database")
            else:
                # Check rate consistency
                if abs(float(item.rate) - float(ssr_match['rate'])) > 0.1:
                    item.validation_errors.append(
                        f"Rate differs from SSR: {ssr_match['rate']}"
                    )
    
    def import_selected_items(self, preview_items: List[ImportPreviewItem], 
                            database_manager, save_as_template: bool = False,
                            template_name: str = None, match_with_ssr: bool = False) -> Dict[str, int]:
        """
        Import selected preview items to database with enhanced features
        
        Args:
            preview_items: List of preview items
            database_manager: Database manager instance
            save_as_template: Whether to save as template
            template_name: Name for template if saving
            match_with_ssr: Whether to match items with SSR database
        
        Returns:
            Dict with import statistics
        """
        selected_items = [item for item in preview_items if item.selected]
        
        if not selected_items:
            return {'imported': 0, 'skipped': 0, 'failed': 0, 'ssr_matched': 0}
        
        # Convert preview items to ScheduleItemModel objects
        schedule_items = []
        ssr_matched_count = 0
        
        for item in selected_items:
            try:
                # Apply SSR matching if requested
                final_rate = item.rate
                final_description = item.description
                
                if match_with_ssr and self.ssr_manager:
                    ssr_item = self.ssr_manager.find_item_by_code(item.code)
                    if not ssr_item:
                        # Try fuzzy matching
                        fuzzy_matches = self.ssr_manager.fuzzy_match_description(
                            item.description, threshold=0.85, limit=1
                        )
                        if fuzzy_matches:
                            ssr_item = fuzzy_matches[0]
                    
                    if ssr_item:
                        final_rate = ssr_item['rate']
                        ssr_matched_count += 1
                        log.info(f"SSR matched item {item.code}: rate updated from {item.rate} to {final_rate}")
                
                schedule_item = schedule.ScheduleItemModel(
                    code=item.code,
                    description=final_description,
                    unit=item.unit,
                    rate=final_rate,
                    qty=item.qty,
                    remarks=item.remarks or '',
                    ana_remarks='',
                    category=item.category,
                    parent=None
                )
                schedule_items.append(schedule_item)
            except Exception as e:
                log.error(f"Error converting preview item {item.code}: {e}")
        
        # Import to database
        imported, skipped, failed = database_manager.insert_schedule_items(schedule_items)
        
        # Save as template if requested
        if save_as_template and template_name and self.template_manager:
            try:
                template_description = f"Template created from Excel import on {misc.get_current_date()}"
                self.template_manager.save_items_as_template(
                    schedule_items, template_name, template_description
                )
                log.info(f"Saved import as template: {template_name}")
            except Exception as e:
                log.error(f"Error saving template: {e}")
        
        return {
            'imported': imported,
            'skipped': skipped, 
            'failed': failed,
            'ssr_matched': ssr_matched_count
        }
    
    def import_with_row_selection(self, filepath: Path, selected_rows: List[int],
                                sheet_name: str = None, column_mapping: Dict[str, int] = None,
                                save_as_template: bool = False, template_name: str = None) -> Dict[str, int]:
        """
        Import specific rows from Excel file
        
        Args:
            filepath: Path to Excel file
            selected_rows: List of row numbers to import
            sheet_name: Name of sheet to import from
            column_mapping: Custom column mapping
            save_as_template: Whether to save as template
            template_name: Name for template
            
        Returns:
            Dict with import statistics
        """
        try:
            # Preview selected rows
            preview_items = self.preview_import(
                filepath, sheet_name, selected_rows=selected_rows
            )
            
            # Mark all items as selected since they were specifically chosen
            for item in preview_items:
                item.selected = True
            
            # Import items
            return self.import_selected_items(
                preview_items, self.database_manager, 
                save_as_template, template_name
            )
            
        except Exception as e:
            log.error(f"Error importing with row selection: {e}")
            return {'imported': 0, 'skipped': 0, 'failed': 0, 'error': str(e)}
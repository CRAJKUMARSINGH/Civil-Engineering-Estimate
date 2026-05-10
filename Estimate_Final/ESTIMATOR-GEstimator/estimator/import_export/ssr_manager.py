#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SSR (Schedule of Rates) Manager for enhanced import/export functionality
"""

import logging
import sqlite3
from pathlib import Path
from decimal import Decimal
from typing import List, Dict, Optional, Tuple
from difflib import SequenceMatcher
import openpyxl

from .. import misc

log = logging.getLogger(__name__)

class SSRMatcher:
    """Enhanced SSR matching with fuzzy logic and intelligent suggestions"""
    
    def __init__(self, ssr_manager):
        """Initialize SSR matcher with SSR manager"""
        self.ssr_manager = ssr_manager
        self.match_confidence = 0.0
    
    def match_imported_items_to_ssr(self, imported_items: List[Dict], 
                                  ssr_database: str = None, threshold: float = 0.85) -> Dict:
        """
        Match imported Excel items with SSR database using intelligent matching
        
        Args:
            imported_items: List of imported item dictionaries
            ssr_database: Optional specific SSR database to use
            threshold: Minimum similarity threshold for fuzzy matching
            
        Returns:
            Dictionary with matched and unmatched items plus statistics
        """
        matched = []
        unmatched = []
        fuzzy_matched = []
        
        for item in imported_items:
            # Try exact code match first
            ssr_item = self.ssr_manager.find_item_by_code(item.get('code', ''))
            
            if ssr_item:
                matched.append({
                    'imported': item,
                    'ssr': ssr_item,
                    'match_type': 'exact_code',
                    'confidence': 1.0
                })
            else:
                # Try fuzzy description match
                fuzzy_matches = self.ssr_manager.fuzzy_match_description(
                    item.get('description', ''), threshold=threshold, limit=1
                )
                
                if fuzzy_matches:
                    best_match = fuzzy_matches[0]
                    fuzzy_matched.append({
                        'imported': item,
                        'ssr': best_match,
                        'match_type': 'fuzzy_description',
                        'confidence': best_match.get('similarity', 0.0)
                    })
                else:
                    unmatched.append(item)
        
        # Generate statistics
        total_items = len(imported_items)
        exact_matches = len(matched)
        fuzzy_matches_count = len(fuzzy_matched)
        unmatched_count = len(unmatched)
        
        return {
            'matched': matched,
            'fuzzy_matched': fuzzy_matched,
            'unmatched': unmatched,
            'statistics': {
                'total_items': total_items,
                'exact_matched': exact_matches,
                'fuzzy_matched': fuzzy_matches_count,
                'unmatched': unmatched_count,
                'match_rate': ((exact_matches + fuzzy_matches_count) / total_items * 100) if total_items > 0 else 0
            }
        }
    
    def suggest_ssr_alternatives(self, item_description: str, limit: int = 5) -> List[Dict]:
        """
        Suggest alternative SSR items for unmatched descriptions
        
        Args:
            item_description: Description to find alternatives for
            limit: Maximum number of suggestions
            
        Returns:
            List of suggested SSR items with similarity scores
        """
        # Use multiple matching strategies
        suggestions = []
        
        # 1. Fuzzy string matching
        fuzzy_matches = self.ssr_manager.fuzzy_match_description(
            item_description, threshold=0.6, limit=limit
        )
        suggestions.extend(fuzzy_matches)
        
        # 2. Keyword-based matching
        keywords = self._extract_keywords(item_description)
        for keyword in keywords:
            keyword_matches = self.ssr_manager.search_items(keyword, limit=2)
            for match in keyword_matches:
                match['match_type'] = 'keyword'
                match['similarity'] = 0.7  # Default similarity for keyword matches
                suggestions.append(match)
        
        # Remove duplicates and sort by similarity
        unique_suggestions = {}
        for suggestion in suggestions:
            code = suggestion.get('code', '')
            if code not in unique_suggestions or suggestion.get('similarity', 0) > unique_suggestions[code].get('similarity', 0):
                unique_suggestions[code] = suggestion
        
        sorted_suggestions = sorted(
            unique_suggestions.values(), 
            key=lambda x: x.get('similarity', 0), 
            reverse=True
        )
        
        return sorted_suggestions[:limit]
    
    def _extract_keywords(self, description: str) -> List[str]:
        """Extract meaningful keywords from description"""
        import re
        
        # Remove common stop words
        stop_words = {
            'of', 'and', 'or', 'the', 'in', 'on', 'at', 'to', 'for', 'with', 'by',
            'from', 'as', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
            'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
            'should', 'may', 'might', 'must', 'can', 'shall'
        }
        
        # Extract words (alphanumeric sequences)
        words = re.findall(r'\b[a-zA-Z0-9]+\b', description.lower())
        
        # Filter out stop words and short words
        keywords = [word for word in words if len(word) > 2 and word not in stop_words]
        
        return keywords[:5]  # Return top 5 keywords

class SSRManager:
    """Manages SSR databases and provides matching functionality"""
    
    def __init__(self, ssr_database_path: Optional[Path] = None):
        """Initialize SSR manager with optional database path"""
        self.ssr_database_path = ssr_database_path
        self.connection = None
        self.ssr_cache = {}  # Cache for frequently accessed items
        
        if ssr_database_path and ssr_database_path.exists():
            self._initialize_database()
    
    def _initialize_database(self):
        """Initialize SSR database connection and create tables if needed"""
        try:
            self.connection = sqlite3.connect(str(self.ssr_database_path))
            self.connection.row_factory = sqlite3.Row  # Enable dict-like access
            
            # Create SSR tables if they don't exist
            self._create_ssr_tables()
            
            log.info(f"SSR database initialized: {self.ssr_database_path}")
            
        except Exception as e:
            log.error(f"Error initializing SSR database: {e}")
            self.connection = None
    
    def _create_ssr_tables(self):
        """Create SSR database tables"""
        cursor = self.connection.cursor()
        
        # SSR Items table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ssr_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                description TEXT NOT NULL,
                unit TEXT NOT NULL,
                rate DECIMAL(15,2) NOT NULL,
                ssr_year INTEGER NOT NULL,
                ssr_type TEXT NOT NULL DEFAULT 'civil',
                category TEXT,
                subcategory TEXT,
                remarks TEXT,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(code, ssr_year, ssr_type)
            )
        ''')
        
        # SSR Categories table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ssr_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                ssr_type TEXT NOT NULL DEFAULT 'civil'
            )
        ''')
        
        # Create indexes for better performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ssr_code ON ssr_items(code)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ssr_desc ON ssr_items(description)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_ssr_year ON ssr_items(ssr_year)')
        
        self.connection.commit()
    
    def import_ssr_from_excel(self, excel_path: Path, ssr_year: int, 
                             ssr_type: str = 'civil') -> Dict[str, int]:
        """
        Import SSR data from Excel file
        
        Args:
            excel_path: Path to Excel file containing SSR data
            ssr_year: Year of the SSR (e.g., 2023)
            ssr_type: Type of SSR ('civil', 'electrical', 'mechanical')
            
        Returns:
            Dict with import statistics
        """
        if not self.connection:
            raise RuntimeError("SSR database not initialized")
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            
            imported = 0
            skipped = 0
            failed = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Parse SSR items from worksheet
                items = self._parse_ssr_worksheet(worksheet, ssr_year, ssr_type)
                
                # Insert items into database
                for item in items:
                    try:
                        result = self._insert_ssr_item(item)
                        if result:
                            imported += 1
                        else:
                            skipped += 1
                    except Exception as e:
                        log.warning(f"Failed to insert SSR item {item.get('code', 'unknown')}: {e}")
                        failed += 1
            
            workbook.close()
            self.connection.commit()
            
            # Clear cache after import
            self.ssr_cache.clear()
            
            log.info(f"SSR import completed: {imported} imported, {skipped} skipped, {failed} failed")
            
            return {
                'imported': imported,
                'skipped': skipped,
                'failed': failed
            }
            
        except Exception as e:
            log.error(f"Error importing SSR from Excel: {e}")
            if self.connection:
                self.connection.rollback()
            raise
    
    def _parse_ssr_worksheet(self, worksheet, ssr_year: int, ssr_type: str) -> List[Dict]:
        """Parse SSR items from worksheet"""
        items = []
        
        # Find header row
        header_row = None
        for row_num in range(1, min(11, worksheet.max_row + 1)):
            row_data = []
            for col_num in range(1, min(10, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = str(cell.value).strip().lower() if cell.value else ''
                row_data.append(value)
            
            # Look for SSR header patterns
            if any('code' in val or 'item' in val for val in row_data) and \
               any('description' in val or 'particular' in val for val in row_data) and \
               any('rate' in val for val in row_data):
                header_row = row_num
                break
        
        if not header_row:
            log.warning(f"Could not find header row in worksheet {worksheet.title}")
            return items
        
        # Map columns
        column_mapping = {}
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=header_row, column=col_num)
            header = str(cell.value).strip().lower() if cell.value else ''
            
            if any(keyword in header for keyword in ['code', 'item no', 'sl.no']):
                column_mapping['code'] = col_num
            elif any(keyword in header for keyword in ['description', 'particular', 'item']):
                column_mapping['description'] = col_num
            elif 'unit' in header:
                column_mapping['unit'] = col_num
            elif 'rate' in header:
                column_mapping['rate'] = col_num
            elif 'category' in header:
                column_mapping['category'] = col_num
        
        # Parse data rows
        for row_num in range(header_row + 1, worksheet.max_row + 1):
            try:
                item = self._parse_ssr_row(worksheet, row_num, column_mapping, 
                                         ssr_year, ssr_type)
                if item:
                    items.append(item)
            except Exception as e:
                log.warning(f"Error parsing SSR row {row_num}: {e}")
        
        return items
    
    def _parse_ssr_row(self, worksheet, row_num: int, column_mapping: Dict[str, int],
                      ssr_year: int, ssr_type: str) -> Optional[Dict]:
        """Parse individual SSR row"""
        
        # Extract values
        code = self._get_cell_value(worksheet, row_num, column_mapping.get('code'))
        description = self._get_cell_value(worksheet, row_num, column_mapping.get('description'))
        unit = self._get_cell_value(worksheet, row_num, column_mapping.get('unit'))
        rate = self._get_cell_value(worksheet, row_num, column_mapping.get('rate'))
        category = self._get_cell_value(worksheet, row_num, column_mapping.get('category'))
        
        # Skip empty rows
        if not code and not description:
            return None
        
        # Validate and format data
        if not code:
            code = f"SSR_{row_num}"
        if not description:
            return None
        if not unit:
            unit = "Unit"
        
        try:
            rate = Decimal(str(rate)) if rate else Decimal('0')
        except:
            rate = Decimal('0')
        
        return {
            'code': str(code).strip(),
            'description': str(description).strip(),
            'unit': str(unit).strip(),
            'rate': rate,
            'ssr_year': ssr_year,
            'ssr_type': ssr_type,
            'category': str(category).strip() if category else None
        }
    
    def _get_cell_value(self, worksheet, row: int, col: Optional[int]):
        """Safely get cell value"""
        if col is None:
            return None
        try:
            return worksheet.cell(row=row, column=col).value
        except:
            return None
    
    def _insert_ssr_item(self, item: Dict) -> bool:
        """Insert SSR item into database"""
        cursor = self.connection.cursor()
        
        try:
            cursor.execute('''
                INSERT OR IGNORE INTO ssr_items 
                (code, description, unit, rate, ssr_year, ssr_type, category)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                item['code'],
                item['description'],
                item['unit'],
                item['rate'],
                item['ssr_year'],
                item['ssr_type'],
                item['category']
            ))
            
            return cursor.rowcount > 0
            
        except Exception as e:
            log.error(f"Error inserting SSR item: {e}")
            return False
    
    def find_item_by_code(self, code: str, ssr_year: int = None) -> Optional[Dict]:
        """Find SSR item by code"""
        if not self.connection:
            return None
        
        # Check cache first
        cache_key = f"{code}_{ssr_year or 'any'}"
        if cache_key in self.ssr_cache:
            return self.ssr_cache[cache_key]
        
        cursor = self.connection.cursor()
        
        if ssr_year:
            cursor.execute('''
                SELECT * FROM ssr_items 
                WHERE code = ? AND ssr_year = ?
                ORDER BY ssr_year DESC
                LIMIT 1
            ''', (code, ssr_year))
        else:
            cursor.execute('''
                SELECT * FROM ssr_items 
                WHERE code = ?
                ORDER BY ssr_year DESC
                LIMIT 1
            ''', (code,))
        
        row = cursor.fetchone()
        if row:
            result = dict(row)
            self.ssr_cache[cache_key] = result
            return result
        
        return None
    
    def fuzzy_match_description(self, description: str, threshold: float = 0.8,
                               ssr_year: int = None, limit: int = 5) -> List[Dict]:
        """
        Find SSR items using fuzzy string matching on description
        
        Args:
            description: Description to match
            threshold: Minimum similarity ratio (0.0 to 1.0)
            ssr_year: Optional year filter
            limit: Maximum number of results
            
        Returns:
            List of matching items with similarity scores
        """
        if not self.connection:
            return []
        
        cursor = self.connection.cursor()
        
        # Get all descriptions for fuzzy matching
        if ssr_year:
            cursor.execute('''
                SELECT * FROM ssr_items 
                WHERE ssr_year = ?
                ORDER BY code
            ''', (ssr_year,))
        else:
            cursor.execute('''
                SELECT * FROM ssr_items 
                ORDER BY ssr_year DESC, code
            ''')
        
        matches = []
        description_lower = description.lower()
        
        for row in cursor.fetchall():
            item = dict(row)
            item_desc_lower = item['description'].lower()
            
            # Calculate similarity ratio
            ratio = SequenceMatcher(None, description_lower, item_desc_lower).ratio()
            
            if ratio >= threshold:
                item['similarity'] = ratio
                matches.append(item)
        
        # Sort by similarity and return top matches
        matches.sort(key=lambda x: x['similarity'], reverse=True)
        return matches[:limit]
    
    def search_items(self, keyword: str, ssr_year: int = None, 
                    category: str = None, limit: int = 50) -> List[Dict]:
        """
        Search SSR items by keyword
        
        Args:
            keyword: Search keyword
            ssr_year: Optional year filter
            category: Optional category filter
            limit: Maximum number of results
            
        Returns:
            List of matching SSR items
        """
        if not self.connection:
            return []
        
        cursor = self.connection.cursor()
        
        # Build query
        query = '''
            SELECT * FROM ssr_items 
            WHERE (description LIKE ? OR code LIKE ?)
        '''
        params = [f'%{keyword}%', f'%{keyword}%']
        
        if ssr_year:
            query += ' AND ssr_year = ?'
            params.append(ssr_year)
        
        if category:
            query += ' AND category = ?'
            params.append(category)
        
        query += ' ORDER BY ssr_year DESC, code LIMIT ?'
        params.append(limit)
        
        cursor.execute(query, params)
        
        return [dict(row) for row in cursor.fetchall()]
    
    def get_categories(self, ssr_type: str = 'civil') -> List[str]:
        """Get list of available categories"""
        if not self.connection:
            return []
        
        cursor = self.connection.cursor()
        cursor.execute('''
            SELECT DISTINCT category FROM ssr_items 
            WHERE category IS NOT NULL AND ssr_type = ?
            ORDER BY category
        ''', (ssr_type,))
        
        return [row[0] for row in cursor.fetchall()]
    
    def get_years(self, ssr_type: str = 'civil') -> List[int]:
        """Get list of available SSR years"""
        if not self.connection:
            return []
        
        cursor = self.connection.cursor()
        cursor.execute('''
            SELECT DISTINCT ssr_year FROM ssr_items 
            WHERE ssr_type = ?
            ORDER BY ssr_year DESC
        ''', (ssr_type,))
        
        return [row[0] for row in cursor.fetchall()]
    
    def get_statistics(self) -> Dict[str, int]:
        """Get SSR database statistics"""
        if not self.connection:
            return {}
        
        cursor = self.connection.cursor()
        
        stats = {}
        
        # Total items
        cursor.execute('SELECT COUNT(*) FROM ssr_items')
        stats['total_items'] = cursor.fetchone()[0]
        
        # Items by year
        cursor.execute('''
            SELECT ssr_year, COUNT(*) FROM ssr_items 
            GROUP BY ssr_year ORDER BY ssr_year DESC
        ''')
        stats['items_by_year'] = dict(cursor.fetchall())
        
        # Items by type
        cursor.execute('''
            SELECT ssr_type, COUNT(*) FROM ssr_items 
            GROUP BY ssr_type
        ''')
        stats['items_by_type'] = dict(cursor.fetchall())
        
        return stats
    
    def close(self):
        """Close database connection"""
        if self.connection:
            self.connection.close()
            self.connection = None
            log.info("SSR database connection closed")
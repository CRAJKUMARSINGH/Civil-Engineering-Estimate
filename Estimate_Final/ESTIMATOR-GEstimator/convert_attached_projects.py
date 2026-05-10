#!/usr/bin/env python3
"""
Excel to .eproj Converter for Attached_Assets/PROJECTS
Converts all .xls* files in Attached_Assets/PROJECTS to .eproj format
Compatible with Windows - outputs to same directory as source files
"""

import sys
import os
from pathlib import Path
from decimal import Decimal
import sqlite3
import glob

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

PROJECT_EXTENSION = '.eproj'
PROJECT_FILE_VER = 'GESTIMATOR_FILE_REFERENCE_VER_2'
MAX_DESC_LEN = 1000

PROJECTS_DIR = "Attached_Assets/PROJECTS"


def log(msg, indent=0):
    """Print message with optional indentation"""
    prefix = "  " * indent
    print(f"{prefix}{msg}")


def create_gestimator_database(output_path, project_name):
    """Create a new GEstimator database with proper schema"""
    
    if output_path.exists():
        output_path.unlink()
    
    conn = sqlite3.connect(str(output_path))
    cursor = conn.cursor()
    
    cursor.execute('PRAGMA foreign_keys=ON;')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ProjectTable (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ScheduleCategoryTable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT UNIQUE,
            "order" INTEGER
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ResourceCategoryTable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description TEXT UNIQUE,
            "order" INTEGER
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ScheduleTable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            description TEXT,
            unit TEXT,
            rate REAL,
            qty REAL,
            remarks TEXT,
            ana_remarks TEXT,
            category_id INTEGER,
            parent_id INTEGER,
            "order" INTEGER,
            suborder INTEGER,
            colour TEXT,
            FOREIGN KEY (category_id) REFERENCES ScheduleCategoryTable(id) ON DELETE CASCADE,
            FOREIGN KEY (parent_id) REFERENCES ScheduleTable(id) ON DELETE CASCADE
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ResourceTable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE,
            description TEXT,
            unit TEXT,
            rate REAL,
            vat REAL,
            discount REAL,
            reference TEXT,
            category_id INTEGER,
            "order" INTEGER,
            FOREIGN KEY (category_id) REFERENCES ResourceCategoryTable(id) ON DELETE CASCADE
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS SequenceTable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_seq INTEGER,
            id_sch INTEGER,
            itemtype INTEGER,
            value REAL,
            code TEXT,
            description TEXT,
            FOREIGN KEY (id_sch) REFERENCES ScheduleTable(id) ON DELETE CASCADE,
            UNIQUE (id_seq, id_sch)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ResourceItemTable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_sch INTEGER,
            id_seq INTEGER,
            id_res INTEGER,
            qty REAL,
            remarks TEXT,
            FOREIGN KEY (id_sch) REFERENCES ScheduleTable(id) ON DELETE CASCADE,
            FOREIGN KEY (id_seq) REFERENCES SequenceTable(id) ON DELETE CASCADE,
            FOREIGN KEY (id_res) REFERENCES ResourceTable(id) ON DELETE CASCADE
        )
    ''')
    
    cursor.execute('INSERT INTO ProjectTable (key, value) VALUES (?, ?)',
                  ('file_version', PROJECT_FILE_VER))
    cursor.execute('INSERT INTO ProjectTable (key, value) VALUES (?, ?)',
                  ('project_name', project_name))
    cursor.execute('INSERT INTO ProjectTable (key, value) VALUES (?, ?)',
                  ('project_item_code', ''))
    cursor.execute('INSERT INTO ProjectTable (key, value) VALUES (?, ?)',
                  ('project_resource_code', ''))
    cursor.execute('INSERT INTO ProjectTable (key, value) VALUES (?, ?)',
                  ('project_measurement', '["Measurement", ["", []]]'))
    
    conn.commit()
    return conn


def parse_excel_project(excel_path):
    """Parse Excel project file and extract schedule data"""
    log(f"Reading Excel file: {excel_path.name}", 1)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    schedule_sheet = None
    
    for sheet_name in wb.sheetnames:
        if 'abs' in sheet_name.lower() and 'gen' not in sheet_name.lower():
            schedule_sheet = wb[sheet_name]
            log(f"Using sheet: {sheet_name}", 1)
            break
    
    if not schedule_sheet:
        for sheet_name in wb.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['schedule', 'boq', 'estimate']):
                schedule_sheet = wb[sheet_name]
                log(f"Using sheet: {sheet_name}", 1)
                break
    
    if not schedule_sheet:
        schedule_sheet = wb.active
        log(f"Using active sheet: {schedule_sheet.title}", 1)
    
    items = []
    header_row = None
    
    for row_num in range(1, min(21, schedule_sheet.max_row + 1)):
        row = list(schedule_sheet.iter_rows(min_row=row_num, max_row=row_num))[0]
        values = [str(cell.value).lower() if cell.value else '' for cell in row]
        
        has_code = any('s.no' in v or 'code' in v or 'item' in v or ('no' in v and 's' in v) for v in values)
        has_desc = any('description' in v or 'particular' in v or 'work' in v for v in values)
        has_qty = any('qty' in v or 'quantity' in v or 'quant' in v for v in values)
        
        if (has_code or has_desc) and has_qty:
            header_row = row_num
            log(f"Found header at row {header_row}", 1)
            break
    
    if not header_row:
        log("Warning: Could not find header row, using row 5", 1)
        header_row = 5
    
    header_cells = list(schedule_sheet.iter_rows(min_row=header_row, max_row=header_row))[0]
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in header_cells]
    
    code_col = None
    desc_col = None
    unit_col = None
    rate_col = None
    qty_col = None
    amount_col = None
    
    for idx, header in enumerate(headers):
        if not code_col and any(k in header for k in ['s.no', 'code', 'item no', 'sl']):
            code_col = idx
        if not desc_col and any(k in header for k in ['description', 'particular', 'work']):
            desc_col = idx
        if not unit_col and 'unit' in header:
            unit_col = idx
        if not rate_col and 'rate' in header:
            rate_col = idx
        if not qty_col and any(k in header for k in ['qty', 'quantity', 'quant']):
            qty_col = idx
        if not amount_col and any(k in header for k in ['amount', 'total']):
            amount_col = idx
    
    for row_num in range(header_row + 1, schedule_sheet.max_row + 1):
        row = list(schedule_sheet.iter_rows(min_row=row_num, max_row=row_num))[0]
        
        try:
            code = row[code_col].value if code_col is not None and code_col < len(row) else None
            description = row[desc_col].value if desc_col is not None and desc_col < len(row) else None
            unit = row[unit_col].value if unit_col is not None and unit_col < len(row) else None
            rate = row[rate_col].value if rate_col is not None and rate_col < len(row) else None
            qty = row[qty_col].value if qty_col is not None and qty_col < len(row) else None
            
            if not code and not description:
                continue
            
            if description:
                desc_lower = str(description).lower()
                if any(k in desc_lower for k in ['description', 'particular']):
                    if not code or str(code).lower() in ['code', 'item', 'sl', 's.no']:
                        continue
                if desc_lower.strip() in ['total', 'sub total', 'grand total'] or desc_lower.startswith('total'):
                    continue
            
            if not code:
                code = f"ITEM_{row_num}"
            if not description:
                continue
            if not unit:
                unit = "Unit"
            
            code = str(code).strip()
            description = str(description).strip()[:MAX_DESC_LEN]
            unit = str(unit).strip()
            
            try:
                rate = float(rate) if rate else 0.0
            except:
                rate = 0.0
            
            try:
                qty = float(qty) if qty else 0.0
            except:
                qty = 0.0
            
            items.append({
                'code': code,
                'description': description,
                'unit': unit,
                'rate': rate,
                'qty': qty
            })
            
        except Exception as e:
            continue
    
    wb.close()
    log(f"Extracted {len(items)} schedule items", 1)
    
    return items


def insert_schedule_items(conn, items):
    """Insert schedule items into database"""
    
    cursor = conn.cursor()
    inserted = 0
    used_codes = set()
    
    for idx, item in enumerate(items, 1):
        original_code = item['code']
        code = original_code
        suffix = 1
        
        while code in used_codes:
            code = f"{original_code}_dup{suffix}"
            suffix += 1
        
        used_codes.add(code)
        
        try:
            cursor.execute('''
                INSERT INTO ScheduleTable 
                (code, description, unit, rate, qty, remarks, ana_remarks, category_id, parent_id, "order", suborder, colour)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (code, item['description'], item['unit'], item['rate'], item['qty'],
                  '', '', None, None, idx, None, None))
            inserted += 1
            if code != original_code:
                log(f"Renamed duplicate code '{original_code}' to '{code}'", 2)
        except Exception as e:
            log(f"Warning: Could not insert item {code}: {e}", 2)
    
    conn.commit()
    log(f"Inserted {inserted} items into database", 1)
    return inserted


def convert_excel_to_eproj(excel_path, output_dir):
    """Convert Excel project file to .eproj format"""
    
    try:
        project_name = excel_path.stem
        output_filename = project_name + PROJECT_EXTENSION
        output_path = output_dir / output_filename
        
        if output_path.exists():
            log(f"SKIPPED: {output_filename} already exists", 1)
            return None
        
        items = parse_excel_project(excel_path)
        
        if not items:
            log(f"ERROR: No schedule items found", 1)
            return False
        
        log(f"Creating .eproj database", 1)
        conn = create_gestimator_database(output_path, project_name)
        
        inserted = insert_schedule_items(conn, items)
        
        conn.close()
        
        log(f"SUCCESS: Created {output_filename} with {inserted} items", 1)
        
        return True
        
    except Exception as e:
        log(f"ERROR: {e}", 1)
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main conversion process"""
    print("="*80)
    print("Excel to .eproj Converter for Attached_Assets/PROJECTS")
    print("="*80)
    print()
    
    projects_dir = Path(PROJECTS_DIR)
    
    if not projects_dir.exists():
        print(f"ERROR: Directory not found: {PROJECTS_DIR}")
        return 1
    
    excel_files = []
    for pattern in ['*.xlsx', '*.xls']:
        excel_files.extend(projects_dir.glob(pattern))
    
    excel_files = [f for f in excel_files if f.is_file()]
    
    if not excel_files:
        print(f"No Excel files found in {PROJECTS_DIR}")
        return 1
    
    print(f"Found {len(excel_files)} Excel file(s):")
    for f in excel_files:
        print(f"  - {f.name}")
    print()
    
    print("="*80)
    print("Starting conversion...")
    print("="*80)
    print()
    
    converted = 0
    skipped = 0
    failed = 0
    
    for excel_file in excel_files:
        print(f"Processing: {excel_file.name}")
        result = convert_excel_to_eproj(excel_file, projects_dir)
        if result is True:
            converted += 1
        elif result is None:
            skipped += 1
        else:
            failed += 1
        print()
    
    print("="*80)
    print("CONVERSION SUMMARY")
    print("="*80)
    print(f"Total files:      {len(excel_files)}")
    print(f"Converted:        {converted}")
    print(f"Skipped:          {skipped}")
    print(f"Failed:           {failed}")
    print()
    print(f"Output directory: {projects_dir.absolute()}")
    print("="*80)
    
    return 0 if failed == 0 else 1


if __name__ == '__main__':
    sys.exit(main())

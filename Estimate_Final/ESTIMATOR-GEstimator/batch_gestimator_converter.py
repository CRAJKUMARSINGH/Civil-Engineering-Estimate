#!/usr/bin/env python3
"""
Batch GEstimator Excel Converter Tool
Converts multiple Excel estimation files to GEstimator-compatible format
"""

import os
import sys
import glob
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class BatchGEstimatorConverter:
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.results = []
        
    def process_directory(self, input_dir, output_dir=None, file_pattern="*.xls*"):
        """Process all Excel files in a directory"""
        if output_dir is None:
            output_dir = os.path.join(input_dir, "gestimator_converted")
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Find all Excel files
        search_pattern = os.path.join(input_dir, file_pattern)
        excel_files = glob.glob(search_pattern)
        
        if not excel_files:
            logger.warning(f"No Excel files found matching pattern: {search_pattern}")
            return
        
        logger.info(f"Found {len(excel_files)} Excel files to process")
        
        for input_file in excel_files:
            try:
                self.process_file(input_file, output_dir)
            except Exception as e:
                logger.error(f"Failed to process {input_file}: {e}")
                self.error_count += 1
                self.results.append({
                    'input_file': input_file,
                    'output_file': None,
                    'status': 'Error',
                    'error': str(e)
                })
    
    def process_file(self, input_file, output_dir):
        """Process a single Excel file"""
        logger.info(f"Processing: {os.path.basename(input_file)}")
        
        # Create output filename
        base_name = Path(input_file).stem
        output_file = os.path.join(output_dir, f"{base_name}_gestimator.xlsx")
        
        try:
            # Analyze and convert
            schedule_data = self.extract_schedule_data(input_file)
            
            if not schedule_data:
                logger.warning(f"No schedule data found in {input_file}")
                self.results.append({
                    'input_file': input_file,
                    'output_file': None,
                    'status': 'No Data',
                    'items_converted': 0
                })
                return
            
            # Create GEstimator format file
            self.create_gestimator_file(schedule_data, output_file)
            
            self.success_count += 1
            self.results.append({
                'input_file': input_file,
                'output_file': output_file,
                'status': 'Success',
                'items_converted': len(schedule_data)
            })
            
            logger.info(f"Successfully converted: {os.path.basename(output_file)} ({len(schedule_data)} items)")
            
        except Exception as e:
            logger.error(f"Error processing {input_file}: {e}")
            self.error_count += 1
            self.results.append({
                'input_file': input_file,
                'output_file': None,
                'status': 'Error',
                'error': str(e)
            })
    
    def extract_schedule_data(self, input_file):
        """Extract schedule data from Excel file"""
        schedule_data = []
        
        try:
            # Try reading with different engines
            engines = ['openpyxl', 'xlrd']
            excel_file = None
            
            for engine in engines:
                try:
                    excel_file = pd.ExcelFile(input_file, engine=engine)
                    break
                except:
                    continue
            
            if excel_file is None:
                raise Exception("Could not read Excel file with any engine")
            
            # Process each sheet
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(input_file, sheet_name=sheet_name, engine=excel_file.engine.name)
                    sheet_data = self.parse_sheet_for_schedule(df, sheet_name)
                    schedule_data.extend(sheet_data)
                except Exception as e:
                    logger.warning(f"Could not process sheet '{sheet_name}' in {input_file}: {e}")
                    continue
            
        except Exception as e:
            logger.error(f"Error extracting data from {input_file}: {e}")
            raise
        
        return schedule_data
    
    def parse_sheet_for_schedule(self, df, sheet_name):
        """Parse a dataframe for schedule data"""
        items = []
        
        # Skip if dataframe is too small
        if df.shape[0] < 2 or df.shape[1] < 3:
            return items
        
        # Look for header patterns
        header_patterns = ['s.no', 'serial', 'item', 'code', 'particulars', 'description']
        data_patterns = ['rate', 'amount', 'quantity', 'qty', 'cost']
        
        # Find potential data rows
        for idx, row in df.iterrows():
            if idx > 50:  # Don't search too far down
                break
                
            row_str = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
            
            # Check if this looks like a header row
            has_header = any(pattern in row_str for pattern in header_patterns)
            has_data_cols = any(pattern in row_str for pattern in data_patterns)
            
            if has_header and has_data_cols:
                # Start extracting data from next row
                return self.extract_data_from_rows(df, idx + 1)
        
        # If no clear header found, try to extract from rows that look like data
        return self.extract_data_heuristically(df)
    
    def extract_data_from_rows(self, df, start_row):
        """Extract data starting from a specific row"""
        items = []
        
        for idx in range(start_row, len(df)):
            row = df.iloc[idx]
            
            # Skip empty rows
            if row.isna().all():
                continue
            
            # Skip rows that don't have a code/number in first few columns
            has_code = False
            for col_idx in range(min(3, len(row))):
                if pd.notna(row.iloc[col_idx]):
                    val = str(row.iloc[col_idx]).strip()
                    if val and (val.isdigit() or len(val) <= 10):
                        has_code = True
                        break
            
            if not has_code:
                continue
            
            # Parse this row as schedule item
            item = self.parse_schedule_row(row)
            if item:
                items.append(item)
        
        return items
    
    def extract_data_heuristically(self, df):
        """Extract data using heuristics when no clear structure is found"""
        items = []
        
        # Look for rows with numeric values that could be quantities/rates
        for idx, row in df.iterrows():
            if idx > 100:  # Don't process too many rows
                break
            
            # Count numeric values in row
            numeric_count = 0
            text_fields = []
            numeric_fields = []
            
            for cell in row:
                if pd.notna(cell):
                    try:
                        num_val = float(str(cell).replace(',', ''))
                        numeric_fields.append(num_val)
                        numeric_count += 1
                    except:
                        if len(str(cell).strip()) > 5:  # Meaningful text
                            text_fields.append(str(cell).strip())
            
            # If row has both text and numbers, it might be a schedule item
            if numeric_count >= 2 and len(text_fields) >= 1:
                item = {
                    'code': str(idx + 1),  # Use row number as code
                    'description': text_fields[0] if text_fields else '',
                    'unit': '',
                    'rate': numeric_fields[0] if len(numeric_fields) > 0 else 0,
                    'qty': numeric_fields[1] if len(numeric_fields) > 1 else 0,
                    'amount': numeric_fields[2] if len(numeric_fields) > 2 else 0,
                    'remarks': ''
                }
                
                # Calculate amount if not provided
                if item['amount'] == 0 and item['rate'] > 0 and item['qty'] > 0:
                    item['amount'] = item['rate'] * item['qty']
                
                items.append(item)
        
        return items
    
    def parse_schedule_row(self, row):
        """Parse a single row as schedule item"""
        row_values = [str(cell) if pd.notna(cell) else '' for cell in row]
        
        # Find the first non-empty value as code
        code = ''
        for val in row_values[:3]:
            if val.strip():
                code = val.strip()
                break
        
        if not code:
            return None
        
        # Find description (usually the longest text field)
        description = ''
        max_len = 0
        for val in row_values:
            if len(val) > max_len and len(val) > 10:
                description = val
                max_len = len(val)
        
        if not description:
            return None
        
        # Extract numeric values
        numeric_values = []
        for val in row_values:
            try:
                num_val = float(str(val).replace(',', ''))
                if num_val > 0:
                    numeric_values.append(num_val)
            except:
                continue
        
        # Create item
        item = {
            'code': code,
            'description': description,
            'unit': '',
            'rate': numeric_values[0] if len(numeric_values) > 0 else 0,
            'qty': numeric_values[1] if len(numeric_values) > 1 else 0,
            'amount': numeric_values[2] if len(numeric_values) > 2 else 0,
            'remarks': ''
        }
        
        # Calculate amount if missing
        if item['amount'] == 0 and item['rate'] > 0 and item['qty'] > 0:
            item['amount'] = item['rate'] * item['qty']
        
        return item
    
    def create_gestimator_file(self, schedule_data, output_file):
        """Create GEstimator-compatible Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Headers
        headers = ['Code', 'Description', 'Unit', 'Rate', 'Qty', 'Amount', 'Remarks']
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add data
        for item in schedule_data:
            row_data = [
                item['code'],
                item['description'],
                item['unit'],
                item['rate'],
                item['qty'],
                item['amount'],
                item['remarks']
            ]
            ws.append(row_data)
        
        # Set column widths
        column_widths = [15, 60, 10, 15, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Save file
        wb.save(output_file)
    
    def print_summary(self):
        """Print conversion summary"""
        print("\n" + "="*60)
        print("BATCH CONVERSION SUMMARY")
        print("="*60)
        print(f"Total files processed: {self.success_count + self.error_count}")
        print(f"Successful conversions: {self.success_count}")
        print(f"Failed conversions: {self.error_count}")
        print()
        
        if self.results:
            print("Detailed Results:")
            print("-" * 60)
            for result in self.results:
                input_name = os.path.basename(result['input_file'])
                if result['status'] == 'Success':
                    output_name = os.path.basename(result['output_file'])
                    print(f"✓ {input_name} → {output_name} ({result['items_converted']} items)")
                elif result['status'] == 'No Data':
                    print(f"⚠ {input_name} → No schedule data found")
                else:
                    print(f"✗ {input_name} → Error: {result.get('error', 'Unknown error')}")

def main():
    parser = argparse.ArgumentParser(description='Batch convert Excel files to GEstimator format')
    parser.add_argument('input_dir', help='Directory containing Excel files')
    parser.add_argument('-o', '--output-dir', help='Output directory (default: input_dir/gestimator_converted)')
    parser.add_argument('-p', '--pattern', default='*.xls*', help='File pattern to match (default: *.xls*)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input_dir):
        print(f"Error: Input directory '{args.input_dir}' not found")
        sys.exit(1)
    
    converter = BatchGEstimatorConverter()
    converter.process_directory(args.input_dir, args.output_dir, args.pattern)
    converter.print_summary()

if __name__ == "__main__":
    main()
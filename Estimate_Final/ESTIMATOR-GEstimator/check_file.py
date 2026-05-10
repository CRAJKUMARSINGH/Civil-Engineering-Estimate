#!/usr/bin/env python3
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('geodesic_dome_gestimator.xlsx')
ws = wb['Schedule']

print("GEstimator Converted File Analysis")
print("=" * 40)
print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# Print headers
headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=1, column=col).value)
print(f"Headers: {headers}")

# Print first 5 data rows
print("\nFirst 5 data rows:")
for row in range(2, min(7, ws.max_row + 1)):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row, column=col).value
        row_data.append(str(cell_value) if cell_value is not None else "")
    print(f"  {row_data}")
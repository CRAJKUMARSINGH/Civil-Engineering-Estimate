"""
Dynamic Excel Template Importer Module

This module provides functionality to discover and load Excel templates from the file system,
extracting comprehensive metadata including formula presence, sheet counts, and named ranges.
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Any
from dataclasses import dataclass
from datetime import datetime

import openpyxl
import xlrd
from openpyxl import load_workbook


@dataclass
class ExcelFileMetadata:
    """
    Metadata container for Excel template files.
    
    Attributes:
        filename: Name of the Excel file
        filepath: Full path to the file
        format: File extension (.xls or .xlsx)
        sheet_count: Number of sheets in the workbook
        last_modified: Last modification timestamp
        file_size: File size in bytes
        has_formulas: Whether the file contains any formulas
        named_ranges: List of named ranges defined in the workbook
    """
    filename: str
    filepath: Path
    format: str
    sheet_count: int
    last_modified: datetime
    file_size: int
    has_formulas: bool
    named_ranges: List[str]


class ExcelTemplateImporter:
    """
    Dynamically imports and processes Excel templates from a specified directory.
    
    This class scans for Excel files, extracts metadata, and maintains a registry
    of discovered templates for use by the template processing system.
    """
    
    def __init__(self, assets_path: str = "Attached_Assets"):
        """
        Initialize the Excel Template Importer.
        
        Args:
            assets_path: Directory path containing Excel template files
        """
        self.assets_path = Path(assets_path)
        self.templates: Dict[str, ExcelFileMetadata] = {}
        self.logger = logging.getLogger(__name__)
        
        # Ensure assets directory exists
        if not self.assets_path.exists():
            self.logger.warning(f"Assets path does not exist: {self.assets_path}")
            self.assets_path.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"Created assets directory: {self.assets_path}")
    
    def scan_for_templates(self) -> List[ExcelFileMetadata]:
        """
        Scan the assets directory for Excel template files.
        
        Discovers all .xls and .xlsx files in the assets directory, extracts
        metadata for each file, and stores them in the template registry.
        
        Returns:
            List of ExcelFileMetadata objects for all discovered templates
        """
        excel_files = []
        
        self.logger.info(f"Scanning for templates in: {self.assets_path}")
        
        # Scan for .xlsx files
        for file_path in self.assets_path.glob("*.xlsx"):
            try:
                metadata = self._extract_metadata(file_path)
                excel_files.append(metadata)
                self.templates[file_path.stem] = metadata
                self.logger.info(f"Loaded template: {file_path.name}")
            except Exception as e:
                self.logger.error(f"Failed to process {file_path}: {e}", exc_info=True)
        
        # Scan for .xls files
        for file_path in self.assets_path.glob("*.xls"):
            # Skip .xlsx files that were already processed
            if file_path.suffix == ".xlsx":
                continue
            try:
                metadata = self._extract_metadata(file_path)
                excel_files.append(metadata)
                self.templates[file_path.stem] = metadata
                self.logger.info(f"Loaded template: {file_path.name}")
            except Exception as e:
                self.logger.error(f"Failed to process {file_path}: {e}", exc_info=True)
        
        self.logger.info(f"Discovered {len(excel_files)} template(s)")
        return excel_files
    
    def _extract_metadata(self, file_path: Path) -> ExcelFileMetadata:
        """
        Extract comprehensive metadata from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            ExcelFileMetadata object containing file information
        """
        # Get file statistics
        stat = file_path.stat()
        
        if file_path.suffix.lower() == '.xlsx':
            return self._extract_metadata_xlsx(file_path, stat)
        else:  # .xls
            return self._extract_metadata_xls(file_path, stat)
    
    def _extract_metadata_xlsx(self, file_path: Path, stat: os.stat_result) -> ExcelFileMetadata:
        """
        Extract metadata from .xlsx file using openpyxl.
        
        Args:
            file_path: Path to the .xlsx file
            stat: File statistics from os.stat()
            
        Returns:
            ExcelFileMetadata object
        """
        wb = load_workbook(file_path, data_only=False, keep_links=False)
        sheet_count = len(wb.sheetnames)
        has_formulas = self._check_for_formulas_xlsx(wb)
        named_ranges = list(wb.defined_names.definedName) if hasattr(wb.defined_names, 'definedName') else []
        named_range_names = [nr.name for nr in named_ranges] if named_ranges else []
        
        wb.close()
        
        return ExcelFileMetadata(
            filename=file_path.name,
            filepath=file_path,
            format=file_path.suffix.lower(),
            sheet_count=sheet_count,
            last_modified=datetime.fromtimestamp(stat.st_mtime),
            file_size=stat.st_size,
            has_formulas=has_formulas,
            named_ranges=named_range_names
        )
    
    def _extract_metadata_xls(self, file_path: Path, stat: os.stat_result) -> ExcelFileMetadata:
        """
        Extract metadata from .xls file using xlrd.
        
        Args:
            file_path: Path to the .xls file
            stat: File statistics from os.stat()
            
        Returns:
            ExcelFileMetadata object
        """
        wb = xlrd.open_workbook(file_path, formatting_info=False)
        sheet_count = wb.nsheets
        has_formulas = self._check_for_formulas_xls(wb)
        # xlrd has limited named range support
        named_ranges = []
        
        return ExcelFileMetadata(
            filename=file_path.name,
            filepath=file_path,
            format=file_path.suffix.lower(),
            sheet_count=sheet_count,
            last_modified=datetime.fromtimestamp(stat.st_mtime),
            file_size=stat.st_size,
            has_formulas=has_formulas,
            named_ranges=named_ranges
        )
    
    def _check_for_formulas_xlsx(self, workbook) -> bool:
        """
        Check if an .xlsx workbook contains any formulas.
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            True if formulas are found, False otherwise
        """
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula cell
                        return True
        return False
    
    def _check_for_formulas_xls(self, workbook) -> bool:
        """
        Check if an .xls workbook contains any formulas.
        
        Args:
            workbook: xlrd Workbook object
            
        Returns:
            True if formulas are found, False otherwise
        """
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    # xlrd cell type 2 is XL_CELL_NUMBER, type 3 is XL_CELL_DATE
                    # Formulas are typically detected through cell.ctype
                    # However, xlrd doesn't directly expose formulas in newer versions
                    # We'll check if the cell has a formula attribute
                    if hasattr(cell, 'formula') and cell.formula:
                        return True
        return False
    
    def reload_template(self, filepath: str):
        """
        Reload a template that has been modified.
        
        Args:
            filepath: Path to the modified template file
        """
        file_path = Path(filepath)
        try:
            metadata = self._extract_metadata(file_path)
            self.templates[file_path.stem] = metadata
            self.logger.info(f"Reloaded template: {file_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to reload template {filepath}: {e}", exc_info=True)
    
    def load_new_template(self, filepath: str):
        """
        Load a newly created template file.
        
        Args:
            filepath: Path to the new template file
        """
        file_path = Path(filepath)
        try:
            metadata = self._extract_metadata(file_path)
            self.templates[file_path.stem] = metadata
            self.logger.info(f"Loaded new template: {file_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to load new template {filepath}: {e}", exc_info=True)

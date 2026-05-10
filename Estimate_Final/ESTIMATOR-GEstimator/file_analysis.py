#!/usr/bin/env python3
"""
Analysis of Excel files that writes directly to a file
"""

import openpyxl
from pathlib import Path

def analyze_excel_file(file_path, output_file):
    """Analyze the structure of an Excel file in detail"""
    output_file.write(f"\nAnalyzing: {file_path.name}\n")
    output_file.write("="*60 + "\n")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        output_file.write(f"Number of sheets: {len(wb.sheetnames)}\n")
        output_file.write(f"Sheet names: {wb.sheetnames}\n")
        
        # Analyze each sheet in detail
        for sheet_name in wb.sheetnames:
            output_file.write(f"\nSheet: {sheet_name}\n")
            sheet = wb[sheet_name]
            
            output_file.write(f"  Dimensions: {sheet.max_row} rows x {sheet.max_column} columns\n")
            
            # Show more rows to find the data
            output_file.write("  First 20 rows:\n")
            for row_num in range(1, min(21, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = str(cell.value)[:30] if cell.value is not None else ''
                    row_data.append(value)
                output_file.write(f"    Row {row_num:2d}: {row_data}\n")
                
            # Try to find where the actual data starts
            output_file.write("\n  Looking for data patterns...\n")
            for row_num in range(1, min(31, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    row_data.append(value)
                
                # Check if this row might contain header or data
                non_empty_count = sum(1 for v in row_data if v is not None and str(v).strip())
                if non_empty_count >= 4:  # At least 4 non-empty cells
                    formatted_row = [str(v)[:20] if v is not None else '' for v in row_data]
                    output_file.write(f"    Row {row_num:2d} (non-empty: {non_empty_count}): {formatted_row}\n")
                    
        wb.close()
        
    except Exception as e:
        output_file.write(f"Error analyzing {file_path.name}: {e}\n")
        import traceback
        traceback.print_exc(file=output_file)

def main():
    """Main analysis function"""
    # Open output file
    with open("file_analysis_output.txt", "w") as f:
        projects_dir = Path("Attached_Assets/PROJECTS")
        excel_files = list(projects_dir.glob("*.xlsx"))
        
        f.write(f"Found {len(excel_files)} Excel files:\n")
        for file_path in excel_files:
            f.write(f"  - {file_path.name}\n")
        
        f.write("\n" + "="*80 + "\n")
        f.write("DETAILED ANALYSIS OF EXCEL FILES\n")
        f.write("="*80 + "\n")
        
        for excel_file in excel_files:
            analyze_excel_file(excel_file, f)
    
    print("Analysis complete. Check file_analysis_output.txt for results.")

if __name__ == '__main__':
    main()
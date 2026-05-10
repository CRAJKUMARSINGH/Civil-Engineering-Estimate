#!/usr/bin/env python3
"""
Fixed Excel to GEstimator Project Converter
Creates databases with exact schema matching GEstimator's expectations
"""

import sys
import os
from pathlib import Path
from decimal import Decimal
import sqlite3

import openpyxl
import appdirs

# GEstimator constants
PROGRAM_NAME = 'GEstimator'
PROGRAM_AUTHOR = 'CPWD'
PROGRAM_VER = '1'
PROJECT_EXTENSION = '.eproj'
PROJECT_FILE_VER = 'GESTIMATOR_FILE_REFERENCE_VER_2'
MAX_DESC_LEN = 1000

PROJECTS_DIR = "Attached_Assets/PROJECTS"

# Write to log file
log_file = open("fixed_conversion.log", "w", encoding="utf-8")

def log(msg):
    """Log to both console and file"""
    print(msg)
    log_file.write(msg + "\n")
    log_file.flush()


def create_gestimator_database(output_path, project_name):
    """Create a new GEstimator database with EXACT schema from reference"""
    log(f"  Creating database: {output_path.name}")
    
    # Create SQLite database
    conn = sqlite3.connect(str(output_path))
    cursor = conn.cursor()
    
    # Enable foreign keys
    cursor.execute('PRAGMA foreign_keys=ON;')
    
    # Create tables with EXACT schema from DSR2021.eproj (lowercase names, correct types)
    
    # projecttable (lowercase!)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projecttable (
            key VARCHAR(255) PRIMARY KEY,
            value VARCHAR(255)
        )
    ''')
    
    # schedulecategorytable (lowercase!)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS schedulecategorytable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description VARCHAR(255) UNIQUE NOT NULL,
            "order" INTEGER NOT NULL
        )
    ''')
    
    # resourcecategorytable (lowercase!)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS resourcecategorytable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            description VARCHAR(255) UNIQUE NOT NULL,
            "order" INTEGER NOT NULL
        )
    ''')
    
    # scheduletable (lowercase!) - EXACT schema
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scheduletable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code VARCHAR(255) UNIQUE NOT NULL,
            description VARCHAR(255) NOT NULL,
            unit VARCHAR(255),
            rate DECIMAL(10, 2),
            qty DECIMAL(10, 5),
            remarks VARCHAR(255),
            ana_remarks VARCHAR(255),
            category_id INTEGER,
            parent_id INTEGER,
            "order" INTEGER NOT NULL,
            suborder INTEGER,
            colour VARCHAR(255),
            FOREIGN KEY (category_id) REFERENCES schedulecategorytable(id) ON DELETE CASCADE,
            FOREIGN KEY (parent_id) REFERENCES scheduletable(id) ON DELETE CASCADE
        )
    ''')
    
    # resourcetable (lowercase!)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS resourcetable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code VARCHAR(255) UNIQUE NOT NULL,
            description VARCHAR(255) NOT NULL,
            unit VARCHAR(255) NOT NULL,
            rate DECIMAL(10, 2) NOT NULL,
            vat DECIMAL(10, 2),
            discount DECIMAL(10, 2),
            reference VARCHAR(255),
            category_id INTEGER,
            "order" INTEGER NOT NULL,
            FOREIGN KEY (category_id) REFERENCES resourcecategorytable(id) ON DELETE CASCADE
        )
    ''')
    
    # sequencetable (lowercase!)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sequencetable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_seq INTEGER,
            id_sch INTEGER,
            itemtype INTEGER,
            value DECIMAL(10, 5),
            code VARCHAR(255),
            description VARCHAR(255),
            FOREIGN KEY (id_sch) REFERENCES scheduletable(id) ON DELETE CASCADE,
            UNIQUE (id_seq, id_sch)
        )
    ''')
    
    # resourceitemtable (lowercase!)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS resourceitemtable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_sch INTEGER,
            id_seq INTEGER,
            id_res INTEGER,
            qty DECIMAL(10, 5),
            remarks VARCHAR(255),
            FOREIGN KEY (id_sch) REFERENCES scheduletable(id) ON DELETE CASCADE,
            FOREIGN KEY (id_seq) REFERENCES sequencetable(id) ON DELETE CASCADE,
            FOREIGN KEY (id_res) REFERENCES resourcetable(id) ON DELETE CASCADE
        )
    ''')
    
    # Insert project settings
    cursor.execute('INSERT INTO projecttable (key, value) VALUES (?, ?)',
                  ('file_version', PROJECT_FILE_VER))
    cursor.execute('INSERT INTO projecttable (key, value) VALUES (?, ?)',
                  ('project_name', project_name))
    cursor.execute('INSERT INTO projecttable (key, value) VALUES (?, ?)',
                  ('project_item_code', ''))
    cursor.execute('INSERT INTO projecttable (key, value) VALUES (?, ?)',
                  ('project_resource_code', ''))
    cursor.execute('INSERT INTO projecttable (key, value) VALUES (?, ?)',
                  ('project_measurement', '["Measurement", ["", []]]'))
    
    conn.commit()
    return conn


def parse_excel_project(excel_path):
    """Parse Excel project file and extract schedule data"""
    log(f"  Parsing Excel file...")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Try to find the schedule sheet - prefer detailed sheets over abstract
    schedule_sheet = None
    
    # First priority: sheets with "ABS" but not "gen" (detailed abstract)
    for sheet_name in wb.sheetnames:
        if 'abs' in sheet_name.lower() and 'gen' not in sheet_name.lower():
            schedule_sheet = wb[sheet_name]
            log(f"  Found schedule sheet: {sheet_name}")
            break
    
    # Second priority: sheets with schedule/boq keywords
    if not schedule_sheet:
        for sheet_name in wb.sheetnames:
            if any(keyword in sheet_name.lower() for keyword in ['schedule', 'boq', 'estimate']):
                schedule_sheet = wb[sheet_name]
                log(f"  Found schedule sheet: {sheet_name}")
                break
    
    # Last resort: use active sheet
    if not schedule_sheet:
        schedule_sheet = wb.active
        log(f"  Using active sheet: {schedule_sheet.title}")
    
    # Parse schedule items
    items = []
    header_row = None
    
    # Find header row (look for common column names)
    for row_num in range(1, min(21, schedule_sheet.max_row + 1)):
        row = list(schedule_sheet.iter_rows(min_row=row_num, max_row=row_num))[0]
        values = [str(cell.value).lower() if cell.value else '' for cell in row]
        
        # Check if this looks like a header row
        has_code = any('s.no' in v or 'code' in v or 'item' in v or ('no' in v and 's' in v) for v in values)
        has_desc = any('description' in v or 'particular' in v or 'work' in v for v in values)
        has_qty = any('qty' in v or 'quantity' in v or 'quant' in v for v in values)
        
        if (has_code or has_desc) and has_qty:
            header_row = row_num
            log(f"  Found header at row {header_row}")
            break
    
    if not header_row:
        log("  Warning: Could not find header row, trying row 5")
        header_row = 5
    
    # Get headers
    header_cells = list(schedule_sheet.iter_rows(min_row=header_row, max_row=header_row))[0]
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in header_cells]
    
    log(f"  Headers found: {[h for h in headers if h][:10]}")
    
    # Find column indices
    code_col = None
    desc_col = None
    unit_col = None
    rate_col = None
    qty_col = None
    
    for idx, header in enumerate(headers):
        if not code_col and any(k in header for k in ['s.no', 'code', 'item no', 'sl']):
            code_col = idx
        if not desc_col and any(k in header for k in ['description', 'particular', 'work']):
            desc_col = idx
        if not unit_col and 'unit' in header:
            unit_col = idx
        if not rate_col and 'rate' in header:
            rate_col = idx
        if not qty_col and any(k in header for k in ['qty', 'quantity', 'quant']):
            qty_col = idx
    
    log(f"  Column mapping: Code={code_col}, Desc={desc_col}, Unit={unit_col}, Rate={rate_col}, Qty={qty_col}")
    
    # Parse data rows
    for row_num in range(header_row + 1, schedule_sheet.max_row + 1):
        row = list(schedule_sheet.iter_rows(min_row=row_num, max_row=row_num))[0]
        
        try:
            # Get values
            code = row[code_col].value if code_col is not None and code_col < len(row) else None
            description = row[desc_col].value if desc_col is not None and desc_col < len(row) else None
            unit = row[unit_col].value if unit_col is not None and unit_col < len(row) else None
            rate = row[rate_col].value if rate_col is not None and rate_col < len(row) else None
            qty = row[qty_col].value if qty_col is not None and qty_col < len(row) else None
            
            # Skip empty rows
            if not code and not description:
                continue
            
            # Skip header-like rows and total rows
            if description:
                desc_lower = str(description).lower()
                if any(k in desc_lower for k in ['description', 'particular']):
                    if not code or str(code).lower() in ['code', 'item', 'sl', 's.no']:
                        continue
                # Skip total rows
                if desc_lower.strip() in ['total', 'sub total', 'grand total'] or desc_lower.startswith('total'):
                    continue
            
            # Validate and convert
            if not code:
                code = f"ITEM_{row_num}"
            if not description:
                continue
            if not unit:
                unit = ""
            
            code = str(code).strip()
            description = str(description).strip()[:MAX_DESC_LEN]
            unit = str(unit).strip()
            
            try:
                rate = float(rate) if rate else 0.0
            except:
                rate = 0.0
            
            try:
                qty = float(qty) if qty else 0.0
            except:
                qty = 0.0
            
            items.append({
                'code': code,
                'description': description,
                'unit': unit,
                'rate': rate,
                'qty': qty
            })
            
        except Exception as e:
            continue
    
    wb.close()
    log(f"  Extracted {len(items)} schedule items")
    
    return items


def insert_schedule_items(conn, items):
    """Insert schedule items into database"""
    log(f"  Inserting {len(items)} items into database...")
    
    cursor = conn.cursor()
    inserted = 0
    
    for idx, item in enumerate(items, 1):
        try:
            cursor.execute('''
                INSERT INTO scheduletable 
                (code, description, unit, rate, qty, remarks, ana_remarks, category_id, parent_id, "order", suborder, colour)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (item['code'], item['description'], item['unit'], item['rate'], item['qty'],
                  '', None, None, None, idx, None, None))
            inserted += 1
        except sqlite3.IntegrityError:
            log(f"    Warning: Duplicate code {item['code']}, skipping")
        except Exception as e:
            log(f"    Warning: Could not insert item {item['code']}: {e}")
    
    conn.commit()
    log(f"  Successfully inserted {inserted} items")
    return inserted


def convert_excel_to_eproj(excel_path, output_dir):
    """Convert Excel project file to .eproj format"""
    log(f"\n{'='*80}")
    log(f"Converting: {excel_path.name}")
    log(f"{'='*80}")
    
    try:
        # Create output filename
        project_name = excel_path.stem
        output_filename = project_name + PROJECT_EXTENSION
        output_path = output_dir / output_filename
        
        # Delete if exists (for testing)
        if output_path.exists():
            output_path.unlink()
            log(f"  Deleted existing file")
        
        # Parse Excel file
        items = parse_excel_project(excel_path)
        
        if not items:
            log(f"  ✗ Error: No schedule items found in Excel file")
            return False
        
        # Create GEstimator database
        conn = create_gestimator_database(output_path, project_name)
        
        # Insert schedule items
        inserted = insert_schedule_items(conn, items)
        
        # Close database
        conn.close()
        
        log(f"  ✓ Successfully created: {output_filename}")
        log(f"  Location: {output_path}")
        log(f"  Items: {inserted}")
        
        # Display project on screen
        log(f"\n  {'─'*76}")
        log(f"  PROJECT CREATED: {project_name}")
        log(f"  {'─'*76}")
        log(f"  File: {output_filename}")
        log(f"  Path: {output_path}")
        log(f"  Schedule Items: {inserted}")
        log(f"  {'─'*76}\n")
        
        return True
        
    except Exception as e:
        log(f"  ✗ Error converting {excel_path.name}: {e}")
        import traceback
        log(traceback.format_exc())
        return False


def main():
    """Main conversion process"""
    log("="*80)
    log("Fixed Excel to GEstimator Project Converter")
    log("="*80)
    
    # Get target directory
    dirs = appdirs.AppDirs(PROGRAM_NAME, PROGRAM_AUTHOR, version=PROGRAM_VER)
    user_data_dir = Path(dirs.user_data_dir)
    target_dir = user_data_dir / 'projects'
    target_dir.mkdir(parents=True, exist_ok=True)
    
    log(f"\nTarget directory: {target_dir}")
    
    # Find Excel project files
    projects_dir = Path(PROJECTS_DIR)
    excel_files = list(projects_dir.glob("*.xlsx"))
    
    log(f"\nFound {len(excel_files)} Excel project files:")
    for f in excel_files:
        log(f"  - {f.name}")
    
    # Convert each file
    log(f"\n{'='*80}")
    log("Starting conversion...")
    log(f"{'='*80}")
    
    converted = 0
    failed = 0
    
    for excel_file in excel_files:
        result = convert_excel_to_eproj(excel_file, target_dir)
        if result is True:
            converted += 1
        else:
            failed += 1
    
    # Summary
    log(f"\n{'='*80}")
    log("CONVERSION SUMMARY")
    log(f"{'='*80}")
    log(f"Total files: {len(excel_files)}")
    log(f"✓ Converted: {converted}")
    log(f"✗ Failed: {failed}")
    log(f"\nConverted projects saved to:")
    log(f"{target_dir}")
    log("="*80)
    
    return 0 if failed == 0 else 1


if __name__ == '__main__':
    try:
        exit_code = main()
        log_file.close()
        sys.exit(exit_code)
    except Exception as e:
        log(f"\nFATAL ERROR: {e}")
        import traceback
        log(traceback.format_exc())
        log_file.close()
        sys.exit(1)

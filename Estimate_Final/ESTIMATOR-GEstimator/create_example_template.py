"""
Create an example Excel template for testing the dynamic template system.
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def create_example_template():
    """Create an example Excel template with input/output cells and formulas."""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Create sheets
    input_sheet = wb.create_sheet("Input")
    calc_sheet = wb.create_sheet("Calc")
    summary_sheet = wb.create_sheet("Summary")
    
    # Remove the default sheet if it exists
    if wb.active is not None:
        wb.remove(wb.active)
    
    # Define styles
    input_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    output_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
    
    # Populate Input sheet
    input_sheet["A1"] = "Length (m)"
    input_sheet["B1"] = 10.0
    input_sheet["B1"].fill = input_fill
    
    input_sheet["A2"] = "Width (m)"
    input_sheet["B2"] = 5.0
    input_sheet["B2"].fill = input_fill
    
    input_sheet["A3"] = "Height (m)"
    input_sheet["B3"] = 3.0
    input_sheet["B3"].fill = input_fill
    
    input_sheet["A5"] = "Material Cost (₹/m³)"
    input_sheet["B5"] = 1500.0
    input_sheet["B5"].fill = input_fill
    
    input_sheet["A6"] = "Labor Cost (₹/m³)"
    input_sheet["B6"] = 750.0
    input_sheet["B6"].fill = input_fill
    
    # Populate Calc sheet
    calc_sheet["A1"] = "Volume Calculation"
    calc_sheet["A2"] = "Length"
    calc_sheet["B2"] = "=Input!B1"
    
    calc_sheet["A3"] = "Width"
    calc_sheet["B3"] = "=Input!B2"
    
    calc_sheet["A4"] = "Height"
    calc_sheet["B4"] = "=Input!B3"
    
    calc_sheet["A5"] = "Volume (m³)"
    calc_sheet["B5"] = "=B2*B3*B4"
    calc_sheet["B5"].fill = output_fill
    
    calc_sheet["A7"] = "Cost Calculation"
    calc_sheet["A8"] = "Material Cost per m³"
    calc_sheet["B8"] = "=Input!B5"
    
    calc_sheet["A9"] = "Labor Cost per m³"
    calc_sheet["B9"] = "=Input!B6"
    
    calc_sheet["A10"] = "Material Cost"
    calc_sheet["B10"] = "=B5*B8"
    calc_sheet["B10"].fill = output_fill
    
    calc_sheet["A11"] = "Labor Cost"
    calc_sheet["B11"] = "=B5*B9"
    calc_sheet["B11"].fill = output_fill
    
    calc_sheet["A12"] = "Total Cost"
    calc_sheet["B12"] = "=B10+B11"
    calc_sheet["B12"].fill = output_fill
    
    # Populate Summary sheet
    summary_sheet["A1"] = "Project Summary"
    summary_sheet["A2"] = "Volume"
    summary_sheet["B2"] = "=Calc!B5"
    summary_sheet["B2"].fill = output_fill
    
    summary_sheet["A3"] = "Total Cost"
    summary_sheet["B3"] = "=Calc!B12"
    summary_sheet["B3"].fill = output_fill
    
    summary_sheet["A5"] = "Cost per m³"
    summary_sheet["B5"] = "=IF(B2<>0,B3/B2,0)"
    summary_sheet["B5"].fill = output_fill
    
    # Save the workbook
    wb.save("Attached_Assets/example_dynamic_template.xlsx")
    print("Example template created: Attached_Assets/example_dynamic_template.xlsx")


if __name__ == "__main__":
    create_example_template()
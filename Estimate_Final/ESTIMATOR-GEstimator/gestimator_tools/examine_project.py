#!/usr/bin/env python3
import openpyxl
from pathlib import Path

project_file = Path("Attached_Assets/PROJECTS/20-40 Construction of Hall with Geodesic Aluminium Dome Roof at Arthuna.xlsx")

print(f"Examining: {project_file.name}")
print("="*80)

try:
    wb = openpyxl.load_workbook(project_file, data_only=True)
    print(f"\nWorkbook sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*80}")
        
        ws = wb[sheet_name]
        
        # Print first 15 rows
        for i, row in enumerate(ws.iter_rows(max_row=15), 1):
            values = [cell.value for cell in row if cell.value is not None]
            if values:
                print(f"Row {i}: {values[:10]}")  # First 10 columns
    
    wb.close()
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

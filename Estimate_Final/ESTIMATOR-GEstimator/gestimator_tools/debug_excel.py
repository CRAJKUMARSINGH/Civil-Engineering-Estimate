#!/usr/bin/env python3
import openpyxl
from pathlib import Path

excel_file = Path("Attached_Assets/PROJECTS/20-40 Construction of Hall with Geodesic Aluminium Dome Roof at Arthuna.xlsx")

log = open("excel_debug.txt", "w", encoding="utf-8")

def write(msg):
    print(msg)
    log.write(msg + "\n")
    log.flush()

write(f"Examining: {excel_file.name}")
write("="*80)

wb = openpyxl.load_workbook(excel_file, data_only=True)
write(f"\nWorkbook sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    write(f"\n{'='*80}")
    write(f"Sheet: {sheet_name}")
    write(f"{'='*80}")
    
    ws = wb[sheet_name]
    write(f"Dimensions: {ws.dimensions}")
    write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
    
    # Print first 30 rows
    for i in range(1, min(31, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=i, max_row=i))[0]
        values = []
        for j, cell in enumerate(row[:15]):  # First 15 columns
            if cell.value is not None:
                val = str(cell.value)[:50]
                values.append(f"[{j}]{val}")
        if values:
            write(f"Row {i:2d}: {' | '.join(values)}")

wb.close()
log.close()
write("\nDone! Check excel_debug.txt")

#!/usr/bin/env python3
"""Quick converter - displays project info"""

import sys
import os

# Write to file immediately
output = open("quick_output.txt", "w", encoding="utf-8")

def log(msg):
    print(msg)
    output.write(msg + "\n")
    output.flush()

log("="*80)
log("Quick Project Converter")
log("="*80)

try:
    log("\n1. Checking Python version...")
    log(f"   Python: {sys.version}")
    
    log("\n2. Checking directories...")
    if os.path.exists("Attached_Assets/PROJECTS"):
        log("   ✓ Projects directory found")
        files = os.listdir("Attached_Assets/PROJECTS")
        log(f"   Found {len(files)} files:")
        for f in files:
            log(f"     - {f}")
    else:
        log("   ✗ Projects directory not found")
    
    log("\n3. Importing openpyxl...")
    import openpyxl
    log("   ✓ openpyxl imported")
    
    log("\n4. Importing peewee...")
    import peewee
    log("   ✓ peewee imported")
    
    log("\n5. Importing GEstimator modules...")
    sys.path.insert(0, '.')
    from estimator import misc
    log(f"   ✓ GEstimator imported - {misc.PROGRAM_NAME}")
    
    log("\n6. Reading first Excel file...")
    excel_file = "Attached_Assets/PROJECTS/20-40 Construction of Hall with Geodesic Aluminium Dome Roof at Arthuna.xlsx"
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    log(f"   ✓ Opened: {os.path.basename(excel_file)}")
    log(f"   Sheets: {wb.sheetnames}")
    
    # Show first sheet structure
    ws = wb.active
    log(f"\n   Active sheet: {ws.title}")
    log(f"   Showing first 15 rows:")
    
    for i, row in enumerate(ws.iter_rows(max_row=15), 1):
        values = []
        for cell in row[:8]:  # First 8 columns
            if cell.value:
                val_str = str(cell.value)[:40]
                values.append(val_str)
        if values:
            log(f"   Row {i:2d}: {' | '.join(values)}")
    
    wb.close()
    
    log("\n" + "="*80)
    log("Analysis complete! Check quick_output.txt for full details.")
    log("="*80)
    
except Exception as e:
    log(f"\nERROR: {e}")
    import traceback
    log(traceback.format_exc())

output.close()
log("Done!")

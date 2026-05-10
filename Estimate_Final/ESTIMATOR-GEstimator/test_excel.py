import openpyxl
from pathlib import Path

# Test reading one of the Excel files
excel_file = Path("Attached_Assets/PROJECTS/20-40 Construction of Hall with Geodesic Aluminium Dome Roof at Arthuna.xlsx")

print(f"Testing file: {excel_file}")
print(f"File exists: {excel_file.exists()}")

if excel_file.exists():
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        print(f"Workbook loaded successfully")
        print(f"Sheet names: {wb.sheetnames}")
        
        # Try to read the first sheet
        sheet = wb.active
        if sheet is not None:
            print(f"Active sheet: {sheet.title}")
            
            # Print first 5 rows
            print("First 5 rows:")
            for row_num, row in enumerate(sheet.iter_rows(max_row=5), 1):
                values = [str(cell.value)[:20] if cell.value else '' for cell in row]
                print(f"  Row {row_num}: {values}")
        else:
            print("No active sheet found")
            
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
else:
    print("File not found!")
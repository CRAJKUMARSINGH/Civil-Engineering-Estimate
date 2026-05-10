#!/usr/bin/env python3
"""
Analyze Excel file structure to understand GEstimator format requirements
"""
import pandas as pd
import xlrd
import openpyxl
import json
import sys
from openpyxl import load_workbook

def analyze_excel_file(filename):
    """Analyze Excel file structure and content"""
    print(f"Analyzing Excel file: {filename}")
    print("="*50)
    
    # Try to load with pandas first to get overview
    try:
        # Get all sheet names
        xls = pd.ExcelFile(filename, engine='xlrd')
        print(f"Total sheets found: {len(xls.sheet_names)}")
        print(f"Sheet names: {xls.sheet_names}")
        print()
        
        # Analyze each sheet
        for sheet_name in xls.sheet_names:
            print(f"Sheet: '{sheet_name}'")
            print("-" * 30)
            
            # Read the sheet
            df = pd.read_excel(filename, sheet_name=sheet_name, engine='xlrd')
            
            print(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            print(f"Columns: {list(df.columns)}")
            
            # Show first few rows
            print("\nFirst 5 rows:")
            print(df.head().to_string())
            
            # Show data types
            print("\nData types:")
            print(df.dtypes)
            
            # Check for empty rows/columns
            empty_rows = df.isnull().all(axis=1).sum()
            empty_cols = df.isnull().all(axis=0).sum()
            print(f"\nEmpty rows: {empty_rows}")
            print(f"Empty columns: {empty_cols}")
            
            print("\n" + "="*50 + "\n")
            
    except Exception as e:
        print(f"Error with pandas/xlrd: {e}")
        
        # Try with openpyxl
        try:
            print("Trying with openpyxl...")
            wb = load_workbook(filename, data_only=True)
            print(f"Worksheet names: {wb.sheetnames}")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\nSheet: '{sheet_name}'")
                print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
                
                # Get some sample data
                for row_num in range(1, min(6, ws.max_row + 1)):
                    row_data = []
                    for col_num in range(1, min(11, ws.max_column + 1)):  # First 10 columns
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    print(f"Row {row_num}: {row_data}")
                    
        except Exception as e2:
            print(f"Error with openpyxl: {e2}")

if __name__ == "__main__":
    analyze_excel_file("project.xls")
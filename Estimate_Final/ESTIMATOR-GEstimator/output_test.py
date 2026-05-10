#!/usr/bin/env python3
"""
Test script that writes output to a file
"""

import openpyxl
from pathlib import Path

def main():
    # Open a file to write output
    with open("test_output.txt", "w") as f:
        f.write("Testing Excel file reading...\n")
        
        # Check if the projects directory exists
        projects_dir = Path("Attached_Assets/PROJECTS")
        f.write(f"Projects directory exists: {projects_dir.exists()}\n")
        
        if projects_dir.exists():
            # List files in the directory
            excel_files = list(projects_dir.glob("*.xlsx"))
            f.write(f"Found {len(excel_files)} Excel files:\n")
            for file_path in excel_files:
                f.write(f"  - {file_path.name}\n")
                
                # Try to open and read the file
                try:
                    f.write(f"    Trying to open {file_path.name}...\n")
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    f.write(f"    Successfully opened {file_path.name}\n")
                    f.write(f"    Sheet names: {wb.sheetnames}\n")
                    
                    # Try to read some data from the first sheet
                    if wb.sheetnames:
                        sheet = wb[wb.sheetnames[0]]
                        f.write(f"    First sheet: {sheet.title}\n")
                        f.write(f"    Max rows: {sheet.max_row}\n")
                        f.write(f"    Max columns: {sheet.max_column}\n")
                        
                        # Read first few rows
                        f.write("    First 5 rows:\n")
                        for row_num in range(1, min(6, sheet.max_row + 1)):
                            row_data = []
                            for col_num in range(1, min(6, sheet.max_column + 1)):
                                cell = sheet.cell(row=row_num, column=col_num)
                                value = str(cell.value)[:20] if cell.value is not None else ''
                                row_data.append(value)
                            f.write(f"      Row {row_num:2d}: {row_data}\n")
                    
                    wb.close()
                    f.write(f"    Closed {file_path.name}\n")
                except Exception as e:
                    f.write(f"    Error reading {file_path.name}: {e}\n")
        else:
            f.write("Projects directory not found!\n")
    
    print("Test completed. Check test_output.txt for results.")

if __name__ == '__main__':
    main()
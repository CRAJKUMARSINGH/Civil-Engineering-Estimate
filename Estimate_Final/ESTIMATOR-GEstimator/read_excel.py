#!/usr/bin/env python3
"""
Simple script to read and display Excel file structure using openpyxl
"""

import openpyxl
from pathlib import Path

def read_excel_file(file_path):
    """Read and display basic info about an Excel file"""
    print(f"\n{'='*60}")
    print(f"Reading file: {file_path.name}")
    print(f"{'='*60}")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        # Read each sheet
        for sheet_name in wb.sheetnames:
            print(f"\nSheet: {sheet_name}")
            sheet = wb[sheet_name]
            
            # Get dimensions
            print(f"Max row: {sheet.max_row}")
            print(f"Max column: {sheet.max_column}")
            
            # Show first 10 rows
            print("First 10 rows:")
            for row_num in range(1, min(11, sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(11, sheet.max_column + 1)):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = str(cell.value)[:30] if cell.value is not None else ''
                    row_data.append(value)
                print(f"  Row {row_num:2d}: {row_data}")
            
            print("\n" + "-"*40)
            
        wb.close()
            
    except Exception as e:
        print(f"Error reading {file_path.name}: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main function"""
    projects_dir = Path("Attached_Assets/PROJECTS")
    excel_files = list(projects_dir.glob("*.xlsx"))
    
    print(f"Found {len(excel_files)} Excel files:")
    for f in excel_files:
        print(f"  - {f.name}")
    
    # Read each file
    for excel_file in excel_files:
        read_excel_file(excel_file)

if __name__ == '__main__':
    main()
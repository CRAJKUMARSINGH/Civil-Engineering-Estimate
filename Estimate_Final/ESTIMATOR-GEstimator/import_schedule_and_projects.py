#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# import_schedule_and_projects.py
#
# Automated import script for GEstimator
# Imports schedule data from Excel files and registers project files
#

import os
import sys
import logging
import sqlite3
import shutil
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime

# Third-party imports
import openpyxl
import peewee
import appdirs

# GEstimator imports
from estimator.data import schedule
from estimator import misc

# Global Constants
SCHEDULES_DIR = "Attached_Assets/SCHEDULES"
PROJECTS_DIR = "Attached_Assets/PROJECTS"
LOG_FILE = "import_log.txt"

# GEstimator constants
PROGRAM_NAME = misc.PROGRAM_NAME
PROGRAM_AUTHOR = misc.PROGRAM_AUTHOR
PROGRAM_VER = misc.PROGRAM_VER
PROJECT_EXTENSION = misc.PROJECT_EXTENSION
MAX_DESC_LEN = misc.MAX_DESC_LEN
PROJECT_FILE_VER = misc.PROJECT_FILE_VER


class ImportLogger:
    """Handles structured logging to file and console"""
    
    def __init__(self, log_filepath):
        """Initialize logger with output file"""
        self.log_filepath = log_filepath
        self.log_entries = []
        self.stats = {
            'excel_files_processed': 0,
            'excel_files_failed': 0,
            'schedule_items_imported': 0,
            'schedule_items_skipped': 0,
            'schedule_items_failed': 0,
            'projects_registered': 0,
            'projects_skipped': 0,
            'projects_failed': 0,
            'total_errors': 0
        }
        
        # Setup file handler
        self.file_handler = open(log_filepath, 'w', encoding='utf-8')
        self._write_header()
    
    def _write_header(self):
        """Write log file header"""
        header = f"""
{'='*80}
GEstimator Schedule and Project Import Log
Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*80}

"""
        self.file_handler.write(header)
        self.file_handler.flush()
    
    def _format_message(self, level, message):
        """Format log message with timestamp"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return f"[{timestamp}] [{level}] {message}"
    
    def info(self, message):
        """Log informational message"""
        formatted = self._format_message('INFO', message)
        print(formatted)
        self.file_handler.write(formatted + '\n')
        self.file_handler.flush()
        self.log_entries.append(('INFO', message))
    
    def warning(self, message):
        """Log warning message"""
        formatted = self._format_message('WARNING', message)
        print(formatted)
        self.file_handler.write(formatted + '\n')
        self.file_handler.flush()
        self.log_entries.append(('WARNING', message))
    
    def error(self, message, exception=None):
        """Log error message with optional exception details"""
        if exception:
            message = f"{message}: {str(exception)}"
        formatted = self._format_message('ERROR', message)
        print(formatted)
        self.file_handler.write(formatted + '\n')
        self.file_handler.flush()
        self.log_entries.append(('ERROR', message))
        self.stats['total_errors'] += 1
    
    def log_file_processing(self, filename, status, details):
        """Log file processing result"""
        message = f"File: {filename} | Status: {status} | {details}"
        if status == 'SUCCESS':
            self.info(message)
        elif status == 'WARNING':
            self.warning(message)
        else:
            self.error(message)
    
    def write_summary(self):
        """Write final summary statistics"""
        summary = f"""
{'='*80}
IMPORT SUMMARY
{'='*80}

Excel Files:
  - Processed: {self.stats['excel_files_processed']}
  - Failed: {self.stats['excel_files_failed']}

Schedule Items:
  - Imported: {self.stats['schedule_items_imported']}
  - Skipped (duplicates): {self.stats['schedule_items_skipped']}
  - Failed: {self.stats['schedule_items_failed']}

Project Files:
  - Registered: {self.stats['projects_registered']}
  - Skipped (already exist): {self.stats['projects_skipped']}
  - Failed: {self.stats['projects_failed']}

Total Errors: {self.stats['total_errors']}

Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*80}
"""
        print(summary)
        self.file_handler.write(summary)
        self.file_handler.flush()
    
    def close(self):
        """Close log file"""
        if self.file_handler:
            self.file_handler.close()


class DatabaseManager:
    """Handles all database operations using GEstimator's ORM models"""
    
    def __init__(self, database_path, logger):
        """Initialize database connection"""
        self.database_path = database_path
        self.logger = logger
        self.database = None
        self.orm_models = None
        self.connected = False
    
    def connect(self):
        """Open database connection and initialize ORM models"""
        try:
            self.logger.info(f"Connecting to database: {self.database_path}")
            
            # Initialize peewee database
            self.database = peewee.SqliteDatabase(self.database_path)
            
            # Get ORM models from GEstimator
            self.orm_models = schedule.get_orm_model(self.database)
            (self.BaseModelSch, self.ProjectTable, self.ScheduleCategoryTable,
             self.ResourceCategoryTable, self.ScheduleTable, self.ResourceTable,
             self.SequenceTable, self.ResourceItemTable, self.Using) = self.orm_models
            
            # Open connection
            self.database.connect()
            
            # Enable foreign key support
            self.database.execute_sql('PRAGMA foreign_keys=ON;')
            
            self.connected = True
            self.logger.info("Database connection established")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to connect to database", e)
            return False
    
    def validate_database(self):
        """Verify database schema and version compatibility"""
        try:
            self.logger.info("Validating database schema...")
            
            # Check if required tables exist
            required_tables = ['ProjectTable', 'ScheduleTable', 'ResourceTable',
                             'ScheduleCategoryTable', 'ResourceCategoryTable']
            
            cursor = self.database.execute_sql(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
            existing_tables = [row[0] for row in cursor.fetchall()]
            
            for table in required_tables:
                if table not in existing_tables:
                    self.logger.error(f"Required table missing: {table}")
                    return False
            
            # Check file version
            try:
                version_query = self.ProjectTable.get(
                    self.ProjectTable.key == 'file_version'
                )
                file_version = version_query.value
                
                if file_version > PROJECT_FILE_VER:
                    self.logger.error(
                        f"Database version {file_version} is newer than supported {PROJECT_FILE_VER}"
                    )
                    return False
                    
                self.logger.info(f"Database version: {file_version}")
                
            except Exception as e:
                self.logger.warning(f"Could not verify file version: {e}")
            
            self.logger.info("Database validation successful")
            return True
            
        except Exception as e:
            self.logger.error(f"Database validation failed", e)
            return False
    
    def check_duplicate(self, code):
        """Check if schedule item with code already exists"""
        try:
            existing = self.ScheduleTable.select().where(
                self.ScheduleTable.code == code
            ).count()
            return existing > 0
        except Exception as e:
            self.logger.error(f"Error checking duplicate for code {code}", e)
            return False
    
    def get_or_create_category(self, category_name):
        """Get existing category or create new one"""
        try:
            if not category_name:
                return None
            
            # Try to get existing category
            try:
                category = self.ScheduleCategoryTable.get(
                    self.ScheduleCategoryTable.description == category_name
                )
                return category
            except self.ScheduleCategoryTable.DoesNotExist:
                # Create new category
                max_order = self.ScheduleCategoryTable.select(
                    peewee.fn.MAX(self.ScheduleCategoryTable.order)
                ).scalar() or 0
                
                category = self.ScheduleCategoryTable.create(
                    description=category_name,
                    order=max_order + 1
                )
                self.logger.info(f"Created new category: {category_name}")
                return category
                
        except Exception as e:
            self.logger.error(f"Error handling category {category_name}", e)
            return None
    
    def insert_schedule_items(self, items):
        """
        Insert schedule items with duplicate checking
        
        Returns:
            tuple: (inserted_count, skipped_count, failed_count)
        """
        inserted = 0
        skipped = 0
        failed = 0
        
        for item in items:
            try:
                # Check for duplicate
                if self.check_duplicate(item.code):
                    self.logger.warning(f"Skipping duplicate item: {item.code}")
                    skipped += 1
                    continue
                
                # Get or create category
                category = None
                if item.category:
                    category = self.get_or_create_category(item.category)
                
                # Get max order for positioning
                max_order = self.ScheduleTable.select(
                    peewee.fn.MAX(self.ScheduleTable.order)
                ).scalar() or 0
                
                # Insert schedule item
                self.ScheduleTable.create(
                    code=item.code,
                    description=item.description,
                    unit=item.unit,
                    rate=item.rate,
                    qty=item.qty,
                    remarks=item.remarks if hasattr(item, 'remarks') else None,
                    ana_remarks=item.ana_remarks if hasattr(item, 'ana_remarks') else None,
                    category=category,
                    parent=None,
                    order=max_order + 1,
                    suborder=None,
                    colour=None
                )
                
                inserted += 1
                self.logger.info(f"Inserted item: {item.code} - {item.description[:50]}")
                
            except Exception as e:
                self.logger.error(f"Failed to insert item {item.code}", e)
                failed += 1
        
        return (inserted, skipped, failed)
    
    def commit_transaction(self):
        """Commit current transaction"""
        try:
            if self.database and self.connected:
                self.database.commit()
                return True
        except Exception as e:
            self.logger.error("Failed to commit transaction", e)
            return False
    
    def rollback_transaction(self):
        """Rollback current transaction"""
        try:
            if self.database and self.connected:
                self.database.rollback()
                self.logger.warning("Transaction rolled back")
                return True
        except Exception as e:
            self.logger.error("Failed to rollback transaction", e)
            return False
    
    def verify_database_integrity(self):
        """Verify database integrity after import"""
        try:
            self.logger.info("Verifying database integrity...")
            
            # Check foreign key constraints
            cursor = self.database.execute_sql('PRAGMA foreign_key_check;')
            violations = cursor.fetchall()
            
            if violations:
                self.logger.error(f"Foreign key constraint violations found: {len(violations)}")
                for violation in violations[:5]:  # Show first 5
                    self.logger.error(f"  Violation: {violation}")
                return False
            
            # Verify all schedule items have valid categories (if set)
            invalid_categories = self.ScheduleTable.select().where(
                (self.ScheduleTable.category.is_null(False)) &
                (~self.ScheduleTable.category.in_(
                    self.ScheduleCategoryTable.select(self.ScheduleCategoryTable.id)
                ))
            ).count()
            
            if invalid_categories > 0:
                self.logger.error(f"Found {invalid_categories} schedule items with invalid categories")
                return False
            
            # Get statistics
            schedule_count = self.ScheduleTable.select().count()
            category_count = self.ScheduleCategoryTable.select().count()
            
            self.logger.info(f"Database integrity verified successfully")
            self.logger.info(f"  Total schedule items: {schedule_count}")
            self.logger.info(f"  Total categories: {category_count}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Database integrity verification failed", e)
            return False
    
    def close(self):
        """Close database connection"""
        try:
            if self.database and self.connected:
                self.database.close()
                self.connected = False
                self.logger.info("Database connection closed")
        except Exception as e:
            self.logger.error("Error closing database", e)

class ExcelScheduleParser:
    """Reads and validates Excel files, converts rows to ScheduleItemModel objects"""
    
    def __init__(self, logger):
        """Initialize parser with logger instance"""
        self.logger = logger
        self.required_columns = ['Code', 'Description', 'Unit', 'Rate', 'Quantity']
    
    def parse_file(self, filepath):
        """
        Parse Excel file and return list of ScheduleItemModel objects
        
        Args:
            filepath: Path to Excel file
            
        Returns:
            tuple: (success: bool, items: list, errors: list)
        """
        items = []
        errors = []
        
        try:
            self.logger.info(f"Parsing Excel file: {filepath}")
            
            # Open workbook
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            worksheet = workbook.active
            
            # Validate headers
            header_valid, header_map = self.validate_headers(worksheet)
            if not header_valid:
                error_msg = f"Invalid headers in {filepath}. Expected: {self.required_columns}"
                self.logger.error(error_msg)
                errors.append(error_msg)
                return (False, items, errors)
            
            # Parse rows (skip header row)
            row_count = 0
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                try:
                    # Check if row is empty
                    if self._is_empty_row(row):
                        continue
                    
                    # Parse row
                    item = self.parse_row(row, row_num, header_map)
                    if item:
                        items.append(item)
                        row_count += 1
                        
                except Exception as e:
                    error_msg = f"Error parsing row {row_num}: {str(e)}"
                    self.logger.warning(error_msg)
                    errors.append(error_msg)
            
            workbook.close()
            
            self.logger.info(f"Parsed {row_count} items from {filepath}")
            return (True, items, errors)
            
        except Exception as e:
            error_msg = f"Failed to parse Excel file {filepath}: {str(e)}"
            self.logger.error(error_msg)
            errors.append(error_msg)
            return (False, items, errors)
    
    def validate_headers(self, worksheet):
        """
        Validate that required columns exist
        
        Returns:
            tuple: (is_valid: bool, header_map: dict)
        """
        try:
            # Get first row as headers
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            headers = [cell.value for cell in header_row if cell.value]
            
            # Create mapping of column names to indices
            header_map = {}
            for idx, header in enumerate(headers):
                if header:
                    header_map[header.strip()] = idx
            
            # Check for required columns
            for required in self.required_columns:
                if required not in header_map:
                    self.logger.error(f"Missing required column: {required}")
                    return (False, {})
            
            return (True, header_map)
            
        except Exception as e:
            self.logger.error(f"Error validating headers: {e}")
            return (False, {})
    
    def parse_row(self, row, row_number, header_map):
        """
        Convert Excel row to ScheduleItemModel
        
        Args:
            row: Excel row object
            row_number: Row number for error reporting
            header_map: Mapping of column names to indices
            
        Returns:
            ScheduleItemModel or None if invalid
        """
        try:
            # Extract values using header map
            code = self._get_cell_value(row, header_map.get('Code'))
            description = self._get_cell_value(row, header_map.get('Description'))
            unit = self._get_cell_value(row, header_map.get('Unit'))
            rate = self._get_cell_value(row, header_map.get('Rate'))
            qty = self._get_cell_value(row, header_map.get('Quantity'))
            
            # Optional fields
            remarks = self._get_cell_value(row, header_map.get('Remarks', -1))
            category = self._get_cell_value(row, header_map.get('Category', -1))
            
            # Validate required fields
            if not code or not description or not unit:
                self.logger.warning(
                    f"Row {row_number}: Missing required field (code, description, or unit)"
                )
                return None
            
            # Convert to string and validate
            code = str(code).strip()
            description = str(description).strip()
            unit = str(unit).strip()
            
            # Validate description length
            if len(description) > MAX_DESC_LEN:
                self.logger.warning(
                    f"Row {row_number}: Description too long ({len(description)} > {MAX_DESC_LEN}), truncating"
                )
                description = description[:MAX_DESC_LEN]
            
            # Convert numeric fields
            try:
                rate = Decimal(str(rate)) if rate else Decimal('0')
            except (InvalidOperation, ValueError) as e:
                self.logger.warning(
                    f"Row {row_number}: Invalid rate value '{rate}', using 0"
                )
                rate = Decimal('0')
            
            try:
                qty = Decimal(str(qty)) if qty else Decimal('0')
            except (InvalidOperation, ValueError) as e:
                self.logger.warning(
                    f"Row {row_number}: Invalid quantity value '{qty}', using 0"
                )
                qty = Decimal('0')
            
            # Create ScheduleItemModel
            item = schedule.ScheduleItemModel(
                code=code,
                description=description,
                unit=unit,
                rate=rate,
                qty=qty,
                remarks=str(remarks) if remarks else '',
                ana_remarks='',
                category=str(category) if category else None,
                parent=None
            )
            
            return item
            
        except Exception as e:
            self.logger.error(f"Row {row_number}: Error parsing row: {e}")
            return None
    
    def _get_cell_value(self, row, col_index):
        """Safely get cell value from row"""
        if col_index is None or col_index < 0 or col_index >= len(row):
            return None
        return row[col_index].value
    
    def _is_empty_row(self, row):
        """Check if all cells in row are empty"""
        return all(cell.value is None or str(cell.value).strip() == '' for cell in row)

class ProjectRegistrar:
    """Validates and registers project files into GEstimator"""
    
    def __init__(self, source_dir, target_dir, logger):
        """Initialize with source and target directories"""
        self.source_dir = Path(source_dir)
        self.target_dir = Path(target_dir)
        self.logger = logger
    
    def discover_projects(self):
        """Find all .eproj files in source directory"""
        try:
            pattern = f"*{PROJECT_EXTENSION}"
            projects = list(self.source_dir.glob(pattern))
            self.logger.info(f"Found {len(projects)} project files in {self.source_dir}")
            return projects
        except Exception as e:
            self.logger.error(f"Error discovering projects", e)
            return []
    
    def validate_project_file(self, filepath):
        """
        Validate that file is a valid GEstimator project
        
        Args:
            filepath: Path to .eproj file
            
        Returns:
            tuple: (is_valid: bool, error_message: str)
        """
        try:
            # Check if file exists
            if not filepath.exists():
                return (False, "File does not exist")
            
            # Try to open as SQLite database
            conn = sqlite3.connect(str(filepath))
            cursor = conn.cursor()
            
            # Check for required tables
            cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )
            tables = [row[0] for row in cursor.fetchall()]
            
            required_tables = ['ProjectTable', 'ScheduleTable', 'ResourceTable']
            missing_tables = [t for t in required_tables if t not in tables]
            
            if missing_tables:
                conn.close()
                return (False, f"Missing required tables: {', '.join(missing_tables)}")
            
            # Check file version
            try:
                cursor.execute(
                    'SELECT value FROM ProjectTable WHERE key="file_version"'
                )
                result = cursor.fetchone()
                if result:
                    file_version = result[0]
                    if file_version > PROJECT_FILE_VER:
                        conn.close()
                        return (False, f"File version {file_version} is newer than supported")
            except Exception as e:
                self.logger.warning(f"Could not verify file version: {e}")
            
            conn.close()
            return (True, "Valid project file")
            
        except sqlite3.DatabaseError as e:
            return (False, f"Invalid SQLite database: {str(e)}")
        except Exception as e:
            return (False, f"Validation error: {str(e)}")
    
    def check_existing(self, filename):
        """Check if project already exists in target directory"""
        target_path = self.target_dir / filename
        return target_path.exists()
    
    def register_project(self, filepath):
        """
        Copy project file to target directory
        
        Args:
            filepath: Source project file path
            
        Returns:
            tuple: (success: bool, target_path: str, message: str)
        """
        try:
            filename = filepath.name
            target_path = self.target_dir / filename
            
            # Check if already exists
            if self.check_existing(filename):
                message = f"Project already exists: {filename}"
                self.logger.warning(message)
                return (False, str(target_path), message)
            
            # Validate project file
            is_valid, validation_msg = self.validate_project_file(filepath)
            if not is_valid:
                message = f"Invalid project file: {validation_msg}"
                self.logger.error(message)
                return (False, None, message)
            
            # Copy file
            shutil.copy2(str(filepath), str(target_path))
            
            message = f"Successfully registered project: {filename}"
            self.logger.info(message)
            return (True, str(target_path), message)
            
        except PermissionError as e:
            message = f"Permission denied: {str(e)}"
            self.logger.error(message)
            return (False, None, message)
        except Exception as e:
            message = f"Failed to register project: {str(e)}"
            self.logger.error(message)
            return (False, None, message)


def initialize_environment(logger):
    """Initialize and validate environment"""
    logger.info("Initializing environment...")
    
    # Validate source directories
    schedules_path = Path(SCHEDULES_DIR)
    projects_path = Path(PROJECTS_DIR)
    
    if not schedules_path.exists():
        raise FileNotFoundError(f"Schedules directory not found: {SCHEDULES_DIR}")
    
    if not projects_path.exists():
        raise FileNotFoundError(f"Projects directory not found: {PROJECTS_DIR}")
    
    logger.info(f"Schedules directory: {schedules_path.absolute()}")
    logger.info(f"Projects directory: {projects_path.absolute()}")
    
    # Resolve GEstimator user data directory
    dirs = appdirs.AppDirs(PROGRAM_NAME, PROGRAM_AUTHOR, version=PROGRAM_VER)
    user_data_dir = Path(dirs.user_data_dir)
    
    logger.info(f"GEstimator user data directory: {user_data_dir}")
    
    # Create target projects directory if needed
    target_projects_dir = user_data_dir / 'projects'
    target_projects_dir.mkdir(parents=True, exist_ok=True)
    logger.info(f"Target projects directory: {target_projects_dir}")
    
    return user_data_dir, target_projects_dir


def run_excel_import(database_manager, parser, logger):
    """Execute Excel import workflow"""
    logger.info("\n" + "="*80)
    logger.info("EXCEL IMPORT PHASE")
    logger.info("="*80)
    
    # Discover Excel files
    schedules_path = Path(SCHEDULES_DIR)
    excel_files = list(schedules_path.glob("*.xlsx"))
    
    logger.info(f"Found {len(excel_files)} Excel files to process")
    
    for excel_file in excel_files:
        logger.info(f"\nProcessing: {excel_file.name}")
        
        try:
            # Parse Excel file
            success, items, errors = parser.parse_file(excel_file)
            
            if not success:
                logger.stats['excel_files_failed'] += 1
                logger.log_file_processing(
                    excel_file.name,
                    'FAILED',
                    f"Parsing failed: {'; '.join(errors[:3])}"
                )
                continue
            
            if not items:
                logger.warning(f"No valid items found in {excel_file.name}")
                logger.stats['excel_files_processed'] += 1
                continue
            
            # Insert items into database
            logger.info(f"Inserting {len(items)} items into database...")
            
            try:
                # Start transaction
                database_manager.database.begin()
                
                inserted, skipped, failed = database_manager.insert_schedule_items(items)
                
                # Commit transaction
                database_manager.commit_transaction()
                
                # Update statistics
                logger.stats['excel_files_processed'] += 1
                logger.stats['schedule_items_imported'] += inserted
                logger.stats['schedule_items_skipped'] += skipped
                logger.stats['schedule_items_failed'] += failed
                
                logger.log_file_processing(
                    excel_file.name,
                    'SUCCESS',
                    f"Imported: {inserted}, Skipped: {skipped}, Failed: {failed}"
                )
                
            except Exception as e:
                # Rollback on error
                database_manager.rollback_transaction()
                logger.stats['excel_files_failed'] += 1
                logger.log_file_processing(
                    excel_file.name,
                    'FAILED',
                    f"Database error: {str(e)}"
                )
                
        except Exception as e:
            logger.stats['excel_files_failed'] += 1
            logger.log_file_processing(
                excel_file.name,
                'FAILED',
                f"Unexpected error: {str(e)}"
            )
    
    logger.info(f"\nExcel import phase completed")


def run_project_registration(registrar, logger):
    """Execute project registration workflow"""
    logger.info("\n" + "="*80)
    logger.info("PROJECT REGISTRATION PHASE")
    logger.info("="*80)
    
    # Discover project files
    projects = registrar.discover_projects()
    
    logger.info(f"Found {len(projects)} project files to register")
    
    for project_file in projects:
        logger.info(f"\nProcessing: {project_file.name}")
        
        try:
            # Register project
            success, target_path, message = registrar.register_project(project_file)
            
            if success:
                logger.stats['projects_registered'] += 1
                logger.log_file_processing(
                    project_file.name,
                    'SUCCESS',
                    message
                )
                # Display project info
                print(f"\n{'='*60}")
                print(f"PROJECT REGISTERED: {project_file.name}")
                print(f"Target Location: {target_path}")
                print(f"{'='*60}\n")
                
            elif "already exists" in message.lower():
                logger.stats['projects_skipped'] += 1
                logger.log_file_processing(
                    project_file.name,
                    'WARNING',
                    message
                )
            else:
                logger.stats['projects_failed'] += 1
                logger.log_file_processing(
                    project_file.name,
                    'FAILED',
                    message
                )
                
        except Exception as e:
            logger.stats['projects_failed'] += 1
            logger.log_file_processing(
                project_file.name,
                'FAILED',
                f"Unexpected error: {str(e)}"
            )
    
    logger.info(f"\nProject registration phase completed")


def main():
    """Main entry point for the import script"""
    logger = None
    database_manager = None
    exit_code = 0
    
    try:
        # Initialize logger
        logger = ImportLogger(LOG_FILE)
        logger.info("=" * 80)
        logger.info("GEstimator Schedule and Project Import Script")
        logger.info("=" * 80)
        
        # Initialize environment
        user_data_dir, target_projects_dir = initialize_environment(logger)
        
        logger.info("\nEnvironment initialized successfully")
        
        # Determine database path
        # Use a temporary database for import, or specify existing database
        # For now, create a new database in user data directory
        db_path = user_data_dir / 'imported_data.eproj'
        
        # Check if database exists, if not create it
        if not db_path.exists():
            logger.info(f"Creating new database: {db_path}")
            # Create new database with GEstimator schema
            temp_db = peewee.SqliteDatabase(str(db_path))
            orm_models = schedule.get_orm_model(temp_db)
            (BaseModelSch, ProjectTable, ScheduleCategoryTable, ResourceCategoryTable,
             ScheduleTable, ResourceTable, SequenceTable, ResourceItemTable, Using) = orm_models
            
            temp_db.connect()
            temp_db.execute_sql('PRAGMA foreign_keys=ON;')
            
            # Create tables
            tables = [ProjectTable, ScheduleTable, ResourceTable,
                     ScheduleCategoryTable, ResourceCategoryTable,
                     SequenceTable, ResourceItemTable]
            temp_db.create_tables(tables)
            
            # Set default project settings
            ProjectTable.create(key='file_version', value=PROJECT_FILE_VER)
            ProjectTable.create(key='project_name', value='Imported Data')
            ProjectTable.create(key='project_item_code', value='')
            ProjectTable.create(key='project_resource_code', value='')
            ProjectTable.create(key='project_measurement', value='["Measurement", ["", []]]')
            
            temp_db.commit()
            temp_db.close()
            logger.info("Database created successfully")
        else:
            logger.info(f"Using existing database: {db_path}")
        
        # Instantiate DatabaseManager
        logger.info("\nConnecting to database...")
        database_manager = DatabaseManager(str(db_path), logger)
        
        if not database_manager.connect():
            raise Exception("Failed to connect to database")
        
        if not database_manager.validate_database():
            raise Exception("Database validation failed")
        
        # Instantiate ExcelScheduleParser
        parser = ExcelScheduleParser(logger)
        
        # Instantiate ProjectRegistrar
        registrar = ProjectRegistrar(PROJECTS_DIR, target_projects_dir, logger)
        
        # Run Excel import
        run_excel_import(database_manager, parser, logger)
        
        # Verify database integrity
        logger.info("\nVerifying database integrity...")
        if not database_manager.verify_database_integrity():
            logger.warning("Database integrity check failed, but continuing...")
        
        # Run project registration
        run_project_registration(registrar, logger)
        
        logger.info("\n" + "="*80)
        logger.info("IMPORT PROCESS COMPLETED SUCCESSFULLY")
        logger.info("="*80)
        
    except FileNotFoundError as e:
        if logger:
            logger.error(f"File or directory not found", e)
        else:
            print(f"ERROR: {e}")
        exit_code = 1
        
    except PermissionError as e:
        if logger:
            logger.error(f"Permission denied", e)
        else:
            print(f"ERROR: Permission denied - {e}")
        exit_code = 1
        
    except Exception as e:
        if logger:
            logger.error(f"Fatal error during import", e)
        else:
            print(f"ERROR: {e}")
            import traceback
            traceback.print_exc()
        exit_code = 1
    
    finally:
        # Cleanup
        if database_manager:
            try:
                database_manager.close()
            except Exception as e:
                if logger:
                    logger.error("Error closing database", e)
        
        if logger:
            logger.write_summary()
            logger.close()
    
    return exit_code


if __name__ == '__main__':
    sys.exit(main())

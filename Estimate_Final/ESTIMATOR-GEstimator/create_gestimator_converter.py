#!/usr/bin/env python3
"""
GEstimator Excel Converter Tool
Converts Excel estimation projects to GEstimator-compatible format

Based on analysis of GEstimator source code:
- Schedule import expects: Code, Description, Unit, Rate, Qty, Amount, Remarks
- Measurements can be imported separately
- Analysis data can be linked to schedule items
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import argparse
import sys
import os

class GEstimatorConverter:
    def __init__(self):
        self.workbook = None
        self.schedule_data = []
        self.measurement_data = []
        
    def analyze_excel_file(self, input_file):
        """Analyze the structure of input Excel file"""
        print(f"Analyzing: {input_file}")
        print("=" * 50)
        
        try:
            excel_file = pd.ExcelFile(input_file, engine='xlrd')
            sheets = excel_file.sheet_names
            print(f"Found {len(sheets)} sheets: {sheets}")
            
            for sheet_name in sheets:
                df = pd.read_excel(input_file, sheet_name=sheet_name, engine='xlrd')
                print(f"\nSheet '{sheet_name}':")
                print(f"  Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
                print(f"  Columns: {list(df.columns)}")
                
                # Look for estimation data patterns
                self._identify_data_patterns(df, sheet_name)
                
        except Exception as e:
            print(f"Error analyzing file: {e}")
    
    def _identify_data_patterns(self, df, sheet_name):
        """Identify common estimation data patterns"""
        # Look for schedule-like data
        potential_schedule_patterns = [
            'particulars', 'description', 'item', 'work',
            'rate', 'amount', 'quantity', 'qty', 'unit'
        ]
        
        # Look for measurement-like data  
        potential_measurement_patterns = [
            'nos', 'length', 'breadth', 'height', 'measurement'
        ]
        
        columns_lower = [str(col).lower() for col in df.columns]
        
        has_schedule_data = any(pattern in ' '.join(columns_lower) 
                              for pattern in potential_schedule_patterns)
        has_measurement_data = any(pattern in ' '.join(columns_lower) 
                                 for pattern in potential_measurement_patterns)
        
        print(f"  Data type indicators:")
        print(f"    Schedule data: {'Yes' if has_schedule_data else 'No'}")
        print(f"    Measurement data: {'Yes' if has_measurement_data else 'No'}")
        
        if has_schedule_data:
            self._extract_schedule_data(df, sheet_name)
        if has_measurement_data:
            self._extract_measurement_data(df, sheet_name)
    
    def _extract_schedule_data(self, df, sheet_name):
        """Extract schedule data from dataframe"""
        print(f"    Extracting schedule data from {sheet_name}...")
        
        # Find data rows (skip headers and empty rows)
        data_start_row = None
        for idx, row in df.iterrows():
            # Look for rows that might contain item data
            row_str = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
            if any(keyword in row_str for keyword in ['s.no', 'particulars', '1', 'item']):
                if any(keyword in row_str for keyword in ['rate', 'amount', 'qty']):
                    data_start_row = idx + 1  # Next row should be data
                    break
        
        if data_start_row is not None and data_start_row < len(df):
            # Extract data rows
            for idx in range(data_start_row, len(df)):
                row = df.iloc[idx]
                if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip():
                    # This looks like a data row
                    item_data = self._parse_schedule_row(row, df.columns)
                    if item_data:
                        self.schedule_data.append(item_data)
    
    def _parse_schedule_row(self, row, columns):
        """Parse a row as schedule item data"""
        # Try to identify columns
        item = {
            'code': '',
            'description': '',
            'unit': '',
            'rate': 0,
            'qty': 0,
            'amount': 0,
            'remarks': ''
        }
        
        row_values = [str(cell) if pd.notna(cell) else '' for cell in row]
        
        # Basic mapping - adjust based on your data structure
        if len(row_values) >= 3:
            # Assume first meaningful column is S.No or Code
            item['code'] = row_values[0].strip()
            
            # Look for description (usually longest text field)
            desc_candidates = [val for val in row_values[1:6] if len(str(val)) > 10]
            if desc_candidates:
                item['description'] = desc_candidates[0]
            
            # Look for numeric values (rate, qty, amount)
            numeric_values = []
            for val in row_values:
                try:
                    num_val = float(str(val).replace(',', ''))
                    if num_val > 0:
                        numeric_values.append(num_val)
                except:
                    continue
            
            if len(numeric_values) >= 2:
                item['qty'] = numeric_values[0] if len(numeric_values) > 0 else 0
                item['rate'] = numeric_values[1] if len(numeric_values) > 1 else 0
                item['amount'] = numeric_values[2] if len(numeric_values) > 2 else item['qty'] * item['rate']
        
        # Only return if we have meaningful data
        if item['code'] and item['description']:
            return item
        return None
    
    def _extract_measurement_data(self, df, sheet_name):
        """Extract measurement data from dataframe"""
        print(f"    Extracting measurement data from {sheet_name}...")
        # Implementation for measurement data extraction
        # This would be similar to schedule extraction but focused on measurements
        pass
    
    def create_gestimator_format(self, output_file):
        """Create GEstimator-compatible Excel file"""
        print(f"\nCreating GEstimator format file: {output_file}")
        
        wb = Workbook()
        
        # Create Schedule sheet
        if self.schedule_data:
            ws_schedule = wb.active
            ws_schedule.title = "Schedule"
            
            # Headers as expected by GEstimator
            headers = ['Code', 'Description', 'Unit', 'Rate', 'Qty', 'Amount', 'Remarks']
            ws_schedule.append(headers)
            
            # Format headers
            for col, header in enumerate(headers, 1):
                cell = ws_schedule.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Add data
            for item in self.schedule_data:
                row_data = [
                    item['code'],
                    item['description'],
                    item['unit'],
                    item['rate'],
                    item['qty'],
                    item['amount'],
                    item['remarks']
                ]
                ws_schedule.append(row_data)
            
            # Set column widths
            column_widths = [15, 50, 10, 15, 15, 15, 20]
            for col, width in enumerate(column_widths, 1):
                ws_schedule.column_dimensions[ws_schedule.cell(row=1, column=col).column_letter].width = width
        
        # Create Measurements sheet if data exists
        if self.measurement_data:
            ws_measurements = wb.create_sheet("Measurements")
            # Add measurement data structure
            # Implementation would go here
        
        # Save the workbook
        wb.save(output_file)
        print(f"GEstimator-compatible file created: {output_file}")
        print(f"Schedule items: {len(self.schedule_data)}")
        
    def convert_file(self, input_file, output_file=None):
        """Convert input file to GEstimator format"""
        if output_file is None:
            base_name = os.path.splitext(input_file)[0]
            output_file = f"{base_name}_gestimator.xlsx"
        
        # Analyze input file
        self.analyze_excel_file(input_file)
        
        # Create GEstimator format
        self.create_gestimator_format(output_file)
        
        return output_file

def main():
    parser = argparse.ArgumentParser(description='Convert Excel estimation files to GEstimator format')
    parser.add_argument('input_file', help='Input Excel file')
    parser.add_argument('-o', '--output', help='Output file name')
    parser.add_argument('-a', '--analyze', action='store_true', help='Only analyze input file structure')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input_file):
        print(f"Error: Input file '{args.input_file}' not found")
        sys.exit(1)
    
    converter = GEstimatorConverter()
    
    if args.analyze:
        converter.analyze_excel_file(args.input_file)
    else:
        output_file = converter.convert_file(args.input_file, args.output)
        print(f"\nConversion completed!")
        print(f"Input: {args.input_file}")
        print(f"Output: {output_file}")
        print(f"\nYou can now import '{output_file}' into GEstimator using:")
        print("Menu -> Schedule Items -> Import from Excel")

if __name__ == "__main__":
    main()
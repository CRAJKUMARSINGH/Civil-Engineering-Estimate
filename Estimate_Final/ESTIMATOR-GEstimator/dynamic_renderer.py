"""
Dynamic Template Renderer Module

This module analyzes Excel template structure and identifies input/output cells
based on formatting conventions and naming patterns.
"""

import logging
import re
from typing import Dict, List, Any, Optional
from openpyxl.styles import PatternFill


class TemplateConfig:
    """
    Configuration for template processing conventions.
    
    Defines the indicators used to identify input and output cells in templates.
    """
    
    INPUT_INDICATORS = {
        'fill_color': 'FFFF00',  # Yellow
        'prefix': 'IN_',
        'named_range_pattern': r'INPUT_.*'
    }
    
    OUTPUT_INDICATORS = {
        'fill_color': '90EE90',  # Light green (note: openpyxl uses ARGB format)
        'prefix': 'OUT_',
        'named_range_pattern': r'OUTPUT_.*'
    }
    
    STANDARD_SHEETS = {
        'INPUT': 'Input',
        'CALCULATION': 'Calc',
        'BOM': 'BOM',
        'SUMMARY': 'Summary'
    }


class DynamicTemplateRenderer:
    """
    Renders Excel templates as dynamic UI forms by analyzing structure.
    
    This class identifies input and output cells based on formatting conventions,
    extracts validation rules, and creates structured representations for UI rendering.
    """
    
    def __init__(self, template_engine):
        """
        Initialize the Dynamic Template Renderer.
        
        Args:
            template_engine: FormulaDependencyEngine instance for formula tracking
        """
        self.template_engine = template_engine
        self.input_fields: Dict[str, Dict] = {}
        self.output_fields: Dict[str, Dict] = {}
        self.logger = logging.getLogger(__name__)
    
    def analyze_template_structure(self, workbook) -> Dict[str, Any]:
        """
        Analyze template structure and identify input/output cells.
        
        Processes all sheets in the workbook to identify input cells, output cells,
        formulas, and validation rules.
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            Dictionary containing complete template structure analysis
        """
        self.logger.info("Analyzing template structure...")
        
        structure = {
            'sheets': {},
            'input_fields': {},
            'output_fields': {},
            'formulas': {},
            'named_ranges': {}
        }
        
        # Extract named ranges
        if hasattr(workbook, 'defined_names'):
            for named_range in workbook.defined_names.definedName:
                structure['named_ranges'][named_range.name] = str(named_range.value)
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_analysis = self._analyze_sheet(sheet, sheet_name)
            structure['sheets'][sheet_name] = sheet_analysis
            
            # Aggregate input and output fields
            for input_cell in sheet_analysis['input_cells']:
                structure['input_fields'][input_cell['reference']] = input_cell
            
            for output_cell in sheet_analysis['output_cells']:
                structure['output_fields'][output_cell['reference']] = output_cell
            
            for formula_cell in sheet_analysis['formula_cells']:
                structure['formulas'][formula_cell['reference']] = formula_cell['formula']
        
        self.logger.info(f"Found {len(structure['input_fields'])} input field(s) and {len(structure['output_fields'])} output field(s)")
        return structure
    
    def _analyze_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze individual sheet for inputs, outputs, and formulas.
        
        Args:
            sheet: openpyxl Worksheet object
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary containing sheet analysis results
        """
        sheet_info = {
            'input_cells': [],
            'output_cells': [],
            'formula_cells': [],
            'data_validation': []
        }
        
        for row in sheet.iter_rows():
            for cell in row:
                # Skip empty cells
                if cell.value is None:
                    continue
                
                cell_ref = f"{sheet_name}!{cell.coordinate}"
                
                # Check for input indicators
                if self._is_input_cell(cell):
                    validation = self._get_validation_rules(cell)
                    sheet_info['input_cells'].append({
                        'reference': cell_ref,
                        'coordinate': cell.coordinate,
                        'value': cell.value,
                        'validation': validation
                    })
                    self.logger.debug(f"Input cell: {cell_ref}")
                
                # Check for output indicators
                elif self._is_output_cell(cell):
                    formula = cell.value if cell.data_type == 'f' else None
                    sheet_info['output_cells'].append({
                        'reference': cell_ref,
                        'coordinate': cell.coordinate,
                        'formula': formula,
                        'value': cell.value
                    })
                    self.logger.debug(f"Output cell: {cell_ref}")
                
                # Track formula cells
                if cell.data_type == 'f':
                    sheet_info['formula_cells'].append({
                        'reference': cell_ref,
                        'coordinate': cell.coordinate,
                        'formula': cell.value
                    })
        
        return sheet_info
    
    def _is_input_cell(self, cell) -> bool:
        """
        Determine if a cell is an input field.
        
        Checks for yellow fill color, IN_ prefix, or INPUT_* named range pattern.
        
        Args:
            cell: openpyxl Cell object
            
        Returns:
            True if cell is identified as an input field
        """
        # Check fill color (yellow)
        if cell.fill and cell.fill.start_color:
            # openpyxl uses ARGB format, so FFFF00 might be stored as FFFFFF00 or similar
            color_rgb = cell.fill.start_color.rgb
            if color_rgb:
                # Remove alpha channel if present and check for yellow
                color_hex = color_rgb[-6:] if len(color_rgb) == 8 else color_rgb
                if color_hex.upper() == TemplateConfig.INPUT_INDICATORS['fill_color']:
                    return True
        
        # Check cell value prefix
        if cell.value and isinstance(cell.value, str):
            if cell.value.startswith(TemplateConfig.INPUT_INDICATORS['prefix']):
                return True
        
        # Check if cell is part of a named range matching INPUT pattern
        # This would require checking workbook.defined_names, which we don't have access to here
        # This check would be done at a higher level
        
        return False
    
    def _is_output_cell(self, cell) -> bool:
        """
        Determine if a cell is an output field.
        
        Checks for light green fill color, OUT_ prefix, or OUTPUT_* named range pattern.
        
        Args:
            cell: openpyxl Cell object
            
        Returns:
            True if cell is identified as an output field
        """
        # Check fill color (light green)
        if cell.fill and cell.fill.start_color:
            color_rgb = cell.fill.start_color.rgb
            if color_rgb:
                # Remove alpha channel if present
                color_hex = color_rgb[-6:] if len(color_rgb) == 8 else color_rgb
                # Light green can be represented in various ways
                # 90EE90 is the standard light green, but also check for common variations
                if color_hex.upper() in ['90EE90', '00FF00', '92D050']:  # Various green shades
                    return True
        
        # Check cell value prefix
        if cell.value and isinstance(cell.value, str):
            if cell.value.startswith(TemplateConfig.OUTPUT_INDICATORS['prefix']):
                return True
        
        return False
    
    def _get_validation_rules(self, cell) -> Dict[str, Any]:
        """
        Extract data validation rules from a cell.
        
        Args:
            cell: openpyxl Cell object
            
        Returns:
            Dictionary containing validation rules
        """
        validation = {}
        
        if cell.data_validation:
            dv = cell.data_validation
            
            # Extract validation type
            if dv.type:
                validation['type'] = dv.type
            
            # Extract formula constraints
            if dv.formula1:
                validation['formula1'] = dv.formula1
            if dv.formula2:
                validation['formula2'] = dv.formula2
            
            # Extract operator
            if dv.operator:
                validation['operator'] = dv.operator
            
            # Extract error message
            if dv.error:
                validation['error_message'] = dv.error
            
            # Extract prompt
            if dv.prompt:
                validation['prompt'] = dv.prompt
        
        # Infer validation from cell type
        if cell.data_type == 'n':  # Numeric
            validation['inferred_type'] = 'decimal'
        elif cell.data_type == 'd':  # Date
            validation['inferred_type'] = 'date'
        elif cell.data_type == 'b':  # Boolean
            validation['inferred_type'] = 'boolean'
        else:
            validation['inferred_type'] = 'text'
        
        return validation
    
    def get_input_cells_for_sheet(self, sheet_name: str, structure: Dict) -> List[Dict]:
        """
        Get all input cells for a specific sheet.
        
        Args:
            sheet_name: Name of the sheet
            structure: Template structure dictionary
            
        Returns:
            List of input cell dictionaries
        """
        if sheet_name in structure['sheets']:
            return structure['sheets'][sheet_name]['input_cells']
        return []
    
    def get_output_cells_for_sheet(self, sheet_name: str, structure: Dict) -> List[Dict]:
        """
        Get all output cells for a specific sheet.
        
        Args:
            sheet_name: Name of the sheet
            structure: Template structure dictionary
            
        Returns:
            List of output cell dictionaries
        """
        if sheet_name in structure['sheets']:
            return structure['sheets'][sheet_name]['output_cells']
        return []

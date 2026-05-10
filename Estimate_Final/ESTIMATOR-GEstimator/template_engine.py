"""
Formula Dependency Engine Module

This module provides functionality to parse Excel formulas, extract dependencies,
and manage calculation order using dependency graph analysis.
"""

import re
import logging
from typing import Dict, Set, Tuple, List, Any
from openpyxl.utils import get_column_letter, column_index_from_string
import networkx as nx


class FormulaDependencyEngine:
    """
    Extracts and manages Excel formula dependencies using graph analysis.
    
    This class parses formulas from Excel workbooks, builds a dependency graph,
    and determines the correct execution order for recalculation.
    """
    
    def __init__(self):
        """Initialize the Formula Dependency Engine."""
        self.dependency_graph = nx.DiGraph()
        self.cell_formulas: Dict[str, str] = {}
        self.cell_values: Dict[str, Any] = {}
        self.logger = logging.getLogger(__name__)
    
    def parse_workbook_formulas(self, workbook) -> nx.DiGraph:
        """
        Parse all formulas in a workbook and build a dependency graph.
        
        Iterates through all sheets and cells, identifying formula cells and
        extracting their dependencies to build a directed graph.
        
        Args:
            workbook: openpyxl Workbook object
            
        Returns:
            NetworkX DiGraph representing formula dependencies
        """
        self.logger.info("Parsing workbook formulas...")
        formula_count = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula cell
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        formula = cell.value
                        
                        if formula:
                            self.cell_formulas[cell_ref] = formula
                            formula_count += 1
                            
                            # Extract dependencies from the formula
                            dependencies = self._extract_dependencies(formula, sheet_name)
                            
                            # Add edges: dependency → dependent
                            for dep in dependencies:
                                self.dependency_graph.add_edge(dep, cell_ref)
                                self.logger.debug(f"Dependency: {dep} → {cell_ref}")
        
        self.logger.info(f"Parsed {formula_count} formula(s) with {self.dependency_graph.number_of_edges()} dependencies")
        return self.dependency_graph
    
    def _extract_dependencies(self, formula: str, current_sheet: str) -> List[str]:
        """
        Extract cell references from a formula string.
        
        Uses regex pattern matching to identify cell references in various formats
        including Sheet!A1 and A1 notation.
        
        Args:
            formula: Formula string to parse
            current_sheet: Name of the sheet containing the formula
            
        Returns:
            List of fully qualified cell references (Sheet!Coordinate)
        """
        dependencies = []
        
        # Pattern to match cell references
        # Matches: Sheet!A1, 'Sheet Name'!A1, A1, $A$1, A$1, $A1
        # Group 1: Optional sheet name with quotes
        # Group 2: Sheet name without quotes
        # Group 3: Cell coordinate with optional $ signs
        cell_pattern = r"(?:'([^']+)'!|([A-Za-z_][A-Za-z0-9_]*)!)?(\$?[A-Z]+\$?\d+)"
        
        matches = re.finditer(cell_pattern, formula, re.IGNORECASE)
        
        for match in matches:
            # Extract sheet name (with or without quotes)
            sheet_with_quotes = match.group(1)
            sheet_without_quotes = match.group(2)
            sheet_name = sheet_with_quotes or sheet_without_quotes or current_sheet
            
            # Extract cell coordinate and remove $ signs
            cell_coord = match.group(3).replace('$', '')
            
            # Create fully qualified reference
            full_ref = f"{sheet_name}!{cell_coord}"
            dependencies.append(full_ref)
        
        return dependencies
    
    def calculate_execution_order(self) -> List[str]:
        """
        Determine the correct calculation order using topological sorting.
        
        Uses the dependency graph to compute an execution order that ensures
        all dependencies are calculated before dependent cells.
        
        Returns:
            List of cell references in execution order
        """
        try:
            execution_order = list(nx.topological_sort(self.dependency_graph))
            self.logger.info(f"Calculated execution order for {len(execution_order)} cells")
            return execution_order
        except nx.NetworkXError as e:
            self.logger.warning(f"Circular dependency detected: {e}")
            return self._handle_circular_dependencies()
    
    def _handle_circular_dependencies(self) -> List[str]:
        """
        Handle circular dependencies in the formula graph.
        
        Identifies strongly connected components (cycles) and returns a
        best-effort execution order.
        
        Returns:
            List of cell references in best-effort execution order
        """
        # Find strongly connected components (cycles)
        cycles = list(nx.strongly_connected_components(self.dependency_graph))
        
        # Log each cycle
        for i, cycle in enumerate(cycles):
            if len(cycle) > 1:
                cycle_refs = ', '.join(cycle)
                self.logger.warning(f"Circular dependency {i+1}: {cycle_refs}")
        
        # Create a condensed graph where each SCC is a single node
        condensed = nx.condensation(self.dependency_graph)
        
        # Get topological order of the condensed graph
        condensed_order = list(nx.topological_sort(condensed))
        
        # Expand back to original nodes
        execution_order = []
        for scc_id in condensed_order:
            # Get the nodes in this strongly connected component
            scc_nodes = condensed.nodes[scc_id]['members']
            execution_order.extend(scc_nodes)
        
        self.logger.info(f"Calculated best-effort execution order for {len(execution_order)} cells with circular dependencies")
        return execution_order
    
    def get_dependents(self, cell_ref: str) -> List[str]:
        """
        Get all cells that depend on the specified cell.
        
        Args:
            cell_ref: Fully qualified cell reference (Sheet!Coordinate)
            
        Returns:
            List of dependent cell references
        """
        if cell_ref in self.dependency_graph:
            return list(self.dependency_graph.successors(cell_ref))
        return []
    
    def get_dependencies(self, cell_ref: str) -> List[str]:
        """
        Get all cells that the specified cell depends on.
        
        Args:
            cell_ref: Fully qualified cell reference (Sheet!Coordinate)
            
        Returns:
            List of dependency cell references
        """
        if cell_ref in self.dependency_graph:
            return list(self.dependency_graph.predecessors(cell_ref))
        return []
    
    def clear(self):
        """Clear all stored formulas and dependency information."""
        self.dependency_graph.clear()
        self.cell_formulas.clear()
        self.cell_values.clear()
        self.logger.debug("Cleared formula dependency engine")

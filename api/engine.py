import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class EngineeringEngine:
    """Ultimate Engineering Engine for Multi-Component Hierarchical Estimates"""
    
    def __init__(self):
        self.settings = {
            'contingency_pct': 5.0,
        }

    def _get_qty_formula(self, unit, row):
        """
        Smart Formula Logic based on Unit of Measurement.
        Detects if the item is Point (Nos/Kg), Linear (m), Area (sqm), or Volume (cum).
        """
        u = str(unit).lower().strip()
        
        # 1. Volume Units (Nos * L * B * D)
        if any(x in u for x in ['cum', 'cu.m', 'cu.mtr', 'cubic']):
            return f"=C{row}*E{row}*G{row}*I{row}"
            
        # 2. Area Units (Nos * L * B)
        if any(x in u for x in ['sqm', 'sq.m', 'sq.mtr', 'square']):
            return f"=C{row}*E{row}*G{row}"
            
        # 3. Linear Units (Nos * L)
        if any(x in u for x in ['rm', 'r.m', 'running', 'mtr', ' m ', 'meter']) or u == 'm':
            # Check carefully to not match 'sqm' or 'cum'
            if not any(x in u for x in ['sq', 'cu']):
                return f"=C{row}*E{row}"
        
        # 4. Weight & Count Units (Direct Value from Nos/Quantity column)
        # Includes: Nos, Each, Set, Job, Kg, Quintal (qum), Tonne, MT
        if any(x in u for x in ['nos', 'each', 'set', 'job', 'kg', 'qtl', 'qum', 'ton', 'mt', 'quintal']):
            return f"=C{row}"
            
        # Default to Volume if unsure
        return f"=C{row}*E{row}*G{row}*I{row}"

    def generate_linked_estimate(self, sections):
        """
        Generates a workbook with one 'gen-abstract' and multiple pairs of ABS/MES sheets.
        """
        wb = openpyxl.Workbook()
        
        # 1. Prepare General Abstract (First Sheet)
        ws_gen = wb.active
        ws_gen.title = "gen-abstract"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center')
        left_align = Alignment(horizontal='left', wrap_text=True)

        ws_gen.append(["S.N.", "Description", "Amount"])
        for cell in ws_gen[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        section_totals_coords = []
        
        # 2. Process each section
        section_idx = 1
        for section_name, items in sections.items():
            # Sanitize sheet names
            safe_name = "".join([c if c.isalnum() or c == ' ' else '_' for c in section_name])[:25]
            abs_title = f"{safe_name}_ABS"
            mes_title = f"{safe_name}_MES"
            
            ws_abs = wb.create_sheet(abs_title)
            ws_meas = wb.create_sheet(mes_title)
            
            # --- MEASUREMENTS SHEET ---
            headers_meas = ["S.N.", "Particulars", "Nos.", "", "Length", "", "Breadth", "", "Height", "", "Qty.", "Units"]
            ws_meas.append(headers_meas)
            for cell in ws_meas[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            qty_coords = {}
            curr_row = 2
            for item in items:
                ws_meas.cell(row=curr_row, column=1, value=item.get('serial'))
                ws_meas.cell(row=curr_row, column=2, value=item.get('description')).alignment = left_align
                ws_meas.cell(row=curr_row, column=3, value=item.get('nos', 1)).alignment = center_align
                ws_meas.cell(row=curr_row, column=4, value="x").alignment = center_align
                ws_meas.cell(row=curr_row, column=5, value=item.get('l', 0)).alignment = center_align
                ws_meas.cell(row=curr_row, column=6, value="x").alignment = center_align
                ws_meas.cell(row=curr_row, column=7, value=item.get('b', 1)).alignment = center_align
                ws_meas.cell(row=curr_row, column=8, value="x").alignment = center_align
                ws_meas.cell(row=curr_row, column=9, value=item.get('d', 1)).alignment = center_align
                ws_meas.cell(row=curr_row, column=10, value="=").alignment = center_align
                
                # SMART QTY FORMULA
                unit = item.get('unit', 'cum')
                formula = self._get_qty_formula(unit, curr_row)
                ws_meas.cell(row=curr_row, column=11, value=formula).alignment = center_align
                ws_meas.cell(row=curr_row, column=12, value=unit).alignment = center_align
                
                qty_coords[item.get('id')] = f"'{mes_title}'!K{curr_row}"
                for col in range(1, 13):
                    ws_meas.cell(row=curr_row, column=col).border = thin_border
                curr_row += 1

            ws_meas.column_dimensions['B'].width = 50
            ws_meas.page_setup.orientation = 'portrait'

            # --- ABSTRACT SHEET ---
            headers_abs = ["S.N.", "Item Description", "Quantity", "Unit", "Rate", "Amount"]
            ws_abs.append(headers_abs)
            for cell in ws_abs[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            curr_row = 2
            for item in items:
                ws_abs.cell(row=curr_row, column=1, value=item.get('serial'))
                desc = f"({item.get('id')}) {item.get('description')}"
                ws_abs.cell(row=curr_row, column=2, value=desc).alignment = left_align
                
                bsr_id = item.get('id')
                if bsr_id in qty_coords:
                    ws_abs.cell(row=curr_row, column=3, value=f"={qty_coords[bsr_id]}")
                else:
                    ws_abs.cell(row=curr_row, column=3, value=item.get('qty', 0))
                
                ws_abs.cell(row=curr_row, column=4, value=item.get('unit', 'cum')).alignment = center_align
                ws_abs.cell(row=curr_row, column=5, value=item.get('rate', 0)).alignment = center_align
                ws_abs.cell(row=curr_row, column=6, value=f"=C{curr_row}*E{curr_row}").alignment = center_align
                
                for col in range(1, 7):
                    ws_abs.cell(row=curr_row, column=col).border = thin_border
                curr_row += 1
            
            total_row = curr_row
            ws_abs.cell(row=total_row, column=5, value="Total:").font = Font(bold=True)
            ws_abs.cell(row=total_row, column=6, value=f"=SUM(F2:F{curr_row-1})").font = Font(bold=True)
            
            section_totals_coords.append(f"'{abs_title}'!F{total_row}")
            ws_abs.column_dimensions['B'].width = 60
            ws_abs.page_setup.orientation = 'portrait'

            # --- ADD TO GENERAL ABSTRACT ---
            gen_row = section_idx + 1
            ws_gen.cell(row=gen_row, column=1, value=section_idx)
            ws_gen.cell(row=gen_row, column=2, value=section_name).alignment = left_align
            ws_gen.cell(row=gen_row, column=3, value=f"={section_totals_coords[-1]}")
            for col in range(1, 4):
                ws_gen.cell(row=gen_row, column=col).border = thin_border
            
            section_idx += 1

        # 3. Finalize General Abstract
        last_section_row = section_idx
        summary_row = last_section_row + 1
        ws_gen.cell(row=summary_row, column=2, value="Sub-Total Amount:").font = Font(bold=True)
        ws_gen.cell(row=summary_row, column=3, value=f"=SUM(C2:C{last_section_row})").font = Font(bold=True)
        
        contingency_row = summary_row + 1
        ws_gen.cell(row=contingency_row, column=2, value=f"Add {self.settings['contingency_pct']}% for Contingencies:").alignment = left_align
        ws_gen.cell(row=contingency_row, column=3, value=f"=C{summary_row}*{self.settings['contingency_pct']/100}")
        
        grand_total_row = contingency_row + 1
        ws_gen.cell(row=grand_total_row, column=2, value="GRAND TOTAL:").font = Font(bold=True)
        ws_gen.cell(row=grand_total_row, column=3, value=f"=C{summary_row}+C{contingency_row}").font = Font(bold=True)
        
        for r in range(summary_row, grand_total_row + 1):
            for c in range(2, 4):
                ws_gen.cell(row=r, column=c).border = thin_border

        ws_gen.column_dimensions['B'].width = 60
        ws_gen.page_setup.orientation = 'portrait'

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream

def generate_linked_estimate(data):
    engine = EngineeringEngine()
    if isinstance(data, list):
        return engine.generate_linked_estimate({"Project Estimate": data})
    else:
        return engine.generate_linked_estimate(data)

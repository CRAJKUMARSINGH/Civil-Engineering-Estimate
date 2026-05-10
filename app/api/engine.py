import openpyxl
from io import BytesIO

def generate_linked_estimate(project_data):
    wb = openpyxl.Workbook()
    
    # Define sheets for General Abstract, Abstract of Cost, and Measurements
    ws_meas = wb.active
    ws_meas.title = "Measurements"
    ws_abs = wb.create_sheet("Abstract")
    ws_gen = wb.create_sheet("General Abstract")

    # --- MEASUREMENTS SHEET ---
    headers_meas = ["Sr. No", "Description", "Nos", "L", "B", "D/H", "Quantity"]
    ws_meas.append(headers_meas)
    qty_coords = {} 

    for i, item in enumerate(project_data, start=2):
        ws_meas.cell(row=i, column=1, value=item['serial'])
        ws_meas.cell(row=i, column=2, value=item['description'])
        ws_meas.cell(row=i, column=3, value=1) # Nos
        ws_meas.cell(row=i, column=4, value=item['qty']) # L
        ws_meas.cell(row=i, column=5, value=1) # B
        ws_meas.cell(row=i, column=6, value=1) # D
        
        # Internal Formula: Qty = Nos * L * B * D
        ws_meas.cell(row=i, column=7, value=f"=C{i}*D{i}*E{i}*F{i}")
        
        # Capture coordinate for Abstract linking
        qty_coords[item['id']] = f"Measurements!G{i}"

    # --- ABSTRACT SHEET ---
    headers_abs = ["Sr. No", "BSR Code", "Description", "Quantity", "Rate", "Amount"]
    ws_abs.append(headers_abs)
    
    for i, item in enumerate(project_data, start=2):
        ws_abs.cell(row=i, column=1, value=item['serial'])
        ws_abs.cell(row=i, column=2, value=item['id'])
        ws_abs.cell(row=i, column=3, value=item['description'])
        
        # AUTO-LINK: Reference the Measurement sheet
        ws_abs.cell(row=i, column=4, value=f"={qty_coords[item['id']]}")
        ws_abs.cell(row=i, column=5, value=item['rate'])
        
        # Amount = (Linked Qty) * Rate
        ws_abs.cell(row=i, column=6, value=f"=D{i}*E{i}")

    # --- GENERAL ABSTRACT ---
    ws_gen.append(["Sl. No", "Description", "Amount"])
    ws_gen.cell(row=2, column=1, value=1)
    ws_gen.cell(row=2, column=2, value="Total as per Abstract")
    last_row_abs = len(project_data) + 1
    ws_gen.cell(row=2, column=3, value=f"=SUM(Abstract!F2:F{last_row_abs})")

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
